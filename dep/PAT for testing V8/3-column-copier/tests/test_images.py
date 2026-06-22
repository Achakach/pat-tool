"""TDD test suite for shift_image_anchors() — RED phase.

All tests fail initially: from src.images import shift_image_anchors raises
ImportError because src/images.py does not exist yet (created in Task 2).

When src/images.py is created, these tests should pass without modification.

AnchorMarker.row is 0-BASED: row 0 = Excel row 1.
"""

import io
import sys
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import (
    OneCellAnchor, TwoCellAnchor, AbsoluteAnchor, AnchorMarker,
)
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D

from src.images import shift_image_anchors  # WILL FAIL until Task 2 creates src/images.py


# ── Constants ──────────────────────────────────────────────────────────────

_PNG_DATA = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
    b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00"
    b"\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82"
)


# ── Test Class ─────────────────────────────────────────────────────────────

class TestShiftImageAnchors:
    """Tests for shift_image_anchors() — TDD RED phase."""

    # ── TC-1 ───────────────────────────────────────────────────────────────

    def test_one_cell_anchor_shifts_below_insert(self):
        """OneCellAnchor at row 7 shifts down when insert_at_row=5, num_rows=3."""
        wb = Workbook()
        ws = wb.active
        img = XlImage(io.BytesIO(_PNG_DATA))
        marker = AnchorMarker(col=0, colOff=0, row=6, rowOff=0)  # row 7
        anchor = OneCellAnchor()
        anchor._from = marker
        anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(1), cy=pixels_to_EMU(1))
        img.anchor = anchor
        ws.add_image(img)

        assert ws._images[0].anchor._from.row == 6, "anchor should start at 0-based row 6"

        shift_image_anchors(ws, insert_at_row=5, num_rows=3)

        assert ws._images[0].anchor._from.row == 9, (
            f"expected _from.row=9, got {ws._images[0].anchor._from.row}"
        )
        wb.close()

    # ── TC-2 ───────────────────────────────────────────────────────────────

    def test_two_cell_anchor_both_shifts_below(self):
        """TwoCellAnchor from row 7→10: both _from and to shift down."""
        wb = Workbook()
        ws = wb.active
        img = XlImage(io.BytesIO(_PNG_DATA))
        from_marker = AnchorMarker(col=0, colOff=0, row=6, rowOff=0)  # row 7
        to_marker = AnchorMarker(col=0, colOff=0, row=9, rowOff=0)    # row 10
        img.anchor = TwoCellAnchor(_from=from_marker, to=to_marker)
        ws.add_image(img)

        assert ws._images[0].anchor._from.row == 6
        assert ws._images[0].anchor.to.row == 9

        shift_image_anchors(ws, insert_at_row=5, num_rows=3)

        assert ws._images[0].anchor._from.row == 9, (
            f"expected _from.row=9, got {ws._images[0].anchor._from.row}"
        )
        assert ws._images[0].anchor.to.row == 12, (
            f"expected to.row=12, got {ws._images[0].anchor.to.row}"
        )
        wb.close()

    # ── TC-3 ───────────────────────────────────────────────────────────────

    def test_two_cell_anchor_spanning_insert(self):
        """TwoCellAnchor spanning insert point: only to.row shifts down."""
        wb = Workbook()
        ws = wb.active
        img = XlImage(io.BytesIO(_PNG_DATA))
        from_marker = AnchorMarker(col=0, colOff=0, row=2, rowOff=0)  # row 3
        to_marker = AnchorMarker(col=0, colOff=0, row=6, rowOff=0)    # row 7
        img.anchor = TwoCellAnchor(_from=from_marker, to=to_marker)
        ws.add_image(img)

        assert ws._images[0].anchor._from.row == 2
        assert ws._images[0].anchor.to.row == 6

        shift_image_anchors(ws, insert_at_row=5, num_rows=3)
        # insert_at_row=5 (0-based: 4). _from.row=2 < 4 → untouched.
        # to.row=6 >= 4 → shifted: 6 + 3 = 9

        assert ws._images[0].anchor._from.row == 2, (
            f"expected _from.row=2 (unchanged), got {ws._images[0].anchor._from.row}"
        )
        assert ws._images[0].anchor.to.row == 9, (
            f"expected to.row=9, got {ws._images[0].anchor.to.row}"
        )
        wb.close()

    # ── TC-4 ───────────────────────────────────────────────────────────────

    def test_absolute_anchor_skipped(self):
        """AbsoluteAnchor has no row attributes — skipped with stderr warning."""
        wb = Workbook()
        ws = wb.active
        img = XlImage(io.BytesIO(_PNG_DATA))
        anchor = AbsoluteAnchor()
        anchor.pos = XDRPoint2D(x=0, y=0)
        anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(1), cy=pixels_to_EMU(1))
        img.anchor = anchor
        ws.add_image(img)

        old_stderr = sys.stderr
        stderr = io.StringIO()
        sys.stderr = stderr
        try:
            shift_image_anchors(ws, insert_at_row=3, num_rows=5)
        finally:
            sys.stderr = old_stderr

        output = stderr.getvalue()
        assert "Skipping AbsoluteAnchor" in output, (
            f"expected 'Skipping AbsoluteAnchor' in stderr, got: {output}"
        )
        wb.close()

    # ── TC-5 ───────────────────────────────────────────────────────────────

    def test_images_above_insert_untouched(self):
        """Image at row 2, insert at row 5 — anchor stays unchanged."""
        wb = Workbook()
        ws = wb.active
        img = XlImage(io.BytesIO(_PNG_DATA))
        marker = AnchorMarker(col=0, colOff=0, row=1, rowOff=0)  # row 2
        anchor = OneCellAnchor()
        anchor._from = marker
        anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(1), cy=pixels_to_EMU(1))
        img.anchor = anchor
        ws.add_image(img)

        assert ws._images[0].anchor._from.row == 1

        shift_image_anchors(ws, insert_at_row=5, num_rows=3)
        # insert_at_row=5 (0-based: 4). _from.row=1 < 4 → no shift

        assert ws._images[0].anchor._from.row == 1, (
            f"expected _from.row=1 (unchanged), got {ws._images[0].anchor._from.row}"
        )
        wb.close()

    # ── TC-6 ───────────────────────────────────────────────────────────────

    def test_no_images_no_crash(self):
        """Worksheet with zero images — function should not raise any exception."""
        wb = Workbook()
        ws = wb.active
        # No calls to ws.add_image() — _images attribute may not even exist

        try:
            shift_image_anchors(ws, insert_at_row=3, num_rows=5)
        except Exception as e:
            pytest.fail(f"shift_image_anchors raised {type(e).__name__}: {e}")

        wb.close()

    # ── TC-7 ───────────────────────────────────────────────────────────────

    def test_zero_rows_noop(self):
        """num_rows=0 — anchor row stays unchanged (no-op)."""
        wb = Workbook()
        ws = wb.active
        img = XlImage(io.BytesIO(_PNG_DATA))
        marker = AnchorMarker(col=0, colOff=0, row=4, rowOff=0)  # row 5
        anchor = OneCellAnchor()
        anchor._from = marker
        anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(1), cy=pixels_to_EMU(1))
        img.anchor = anchor
        ws.add_image(img)

        assert ws._images[0].anchor._from.row == 4

        shift_image_anchors(ws, insert_at_row=3, num_rows=0)

        assert ws._images[0].anchor._from.row == 4, (
            f"expected _from.row=4 (unchanged, num_rows=0), got {ws._images[0].anchor._from.row}"
        )
        wb.close()

    # ── TC-8 ───────────────────────────────────────────────────────────────

    def test_multiple_images_all_shift(self):
        """3 OneCellAnchor images all below insert point — all shift correctly."""
        wb = Workbook()
        ws = wb.active

        def _make_img(row_0based):
            img = XlImage(io.BytesIO(_PNG_DATA))
            marker = AnchorMarker(col=0, colOff=0, row=row_0based, rowOff=0)
            anchor = OneCellAnchor()
            anchor._from = marker
            anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(1), cy=pixels_to_EMU(1))
            img.anchor = anchor
            return img

        img1 = _make_img(7)   # row 8  → 0-based row 7
        img2 = _make_img(9)   # row 10 → 0-based row 9
        img3 = _make_img(11)  # row 12 → 0-based row 11

        ws.add_image(img1)
        ws.add_image(img2)
        ws.add_image(img3)

        assert ws._images[0].anchor._from.row == 7
        assert ws._images[1].anchor._from.row == 9
        assert ws._images[2].anchor._from.row == 11

        shift_image_anchors(ws, insert_at_row=5, num_rows=4)
        # insert_at_row=5 (0-based: 4). All >= 4 → shift by +4

        assert ws._images[0].anchor._from.row == 11, (
            f"img1: expected _from.row=11, got {ws._images[0].anchor._from.row}"
        )
        assert ws._images[1].anchor._from.row == 13, (
            f"img2: expected _from.row=13, got {ws._images[1].anchor._from.row}"
        )
        assert ws._images[2].anchor._from.row == 15, (
            f"img3: expected _from.row=15, got {ws._images[2].anchor._from.row}"
        )
        wb.close()

    # ── TC-9 ───────────────────────────────────────────────────────────────

    def test_survives_save_reload(self):
        """Shifted anchors persist through save-to-BytesIO and reload."""
        wb = Workbook()
        ws = wb.active
        img = XlImage(io.BytesIO(_PNG_DATA))
        marker = AnchorMarker(col=0, colOff=0, row=6, rowOff=0)  # row 7
        anchor = OneCellAnchor()
        anchor._from = marker
        anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(1), cy=pixels_to_EMU(1))
        img.anchor = anchor
        ws.add_image(img)

        assert ws._images[0].anchor._from.row == 6

        shift_image_anchors(ws, insert_at_row=3, num_rows=5)
        # insert_at_row=3 (0-based: 2). _from.row=6 >= 2 → 6 + 5 = 11

        assert ws._images[0].anchor._from.row == 11

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        buf.seek(0)
        wb2 = load_workbook(buf)
        ws2 = wb2.active

        assert ws2._images[0].anchor._from.row == 11, (
            f"after save/reload: expected _from.row=11, got {ws2._images[0].anchor._from.row}"
        )
        wb2.close()

    # ── TC-10 ──────────────────────────────────────────────────────────────

    def test_image_exactly_at_insert_row_shifts(self):
        """Image at insert row (row 5), num_rows=2 — shifts from row 4 to row 6."""
        wb = Workbook()
        ws = wb.active
        img = XlImage(io.BytesIO(_PNG_DATA))
        marker = AnchorMarker(col=0, colOff=0, row=4, rowOff=0)  # row 5
        anchor = OneCellAnchor()
        anchor._from = marker
        anchor.ext = XDRPositiveSize2D(cx=pixels_to_EMU(1), cy=pixels_to_EMU(1))
        img.anchor = anchor
        ws.add_image(img)

        assert ws._images[0].anchor._from.row == 4

        shift_image_anchors(ws, insert_at_row=5, num_rows=2)
        # insert_at_row=5 (0-based: 4). _from.row=4 >= 4 → 4 + 2 = 6

        assert ws._images[0].anchor._from.row == 6, (
            f"expected _from.row=6, got {ws._images[0].anchor._from.row}"
        )
        wb.close()
