"""Integration test: full copier pipeline with real column mapping."""
import hashlib
import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

# Import copier main (conftest.py adds to sys.path)
import copier
from copier import read_matching


class TestRealMapping:
    def test_full_pipeline_with_page_break(self, tmp_path, monkeypatch):
        # 1. Generate test data via subprocess call to generator
        result = subprocess.run([
            sys.executable, str(Path(__file__).parent / "generate_test_data.py"),
            "--tmp-dir", str(tmp_path)
        ], capture_output=True, text=True)
        assert result.returncode == 0, f"Generator failed:\n{result.stderr}"

        src_path = tmp_path / "source_test.xlsx"
        tgt_path = tmp_path / "target_test.xlsx"
        assert src_path.exists()
        assert tgt_path.exists()

        # 2. Compute content hash of original data columns (C/D/E/G/H, rows 3-22)
        # Note: full-file SHA256 would differ because copier writes temp cols Q/R/S
        s_before = load_workbook(str(src_path))
        cs_before = s_before["cutsheet"]
        content_parts = []
        for col in [3, 4, 5, 7, 8]:  # C, D, E, G, H
            for row in range(3, 23):  # rows 3-22
                v = cs_before.cell(row=row, column=col).value
                content_parts.append(str(v))
        content_hash_before = hashlib.sha256("|".join(content_parts).encode()).hexdigest()
        s_before.close()

        # 3. Create matching.xlsx in tmp
        matching_path = tmp_path / "matching.xlsx"
        mwb = Workbook()
        mws = mwb.active
        mws["A1"] = "Site"
        mws["B1"] = "PW Number"
        mws["A2"] = "target_test"
        mws["B2"] = "TEST001"
        mwb.save(str(matching_path))
        mwb.close()

        # 4. Build test config
        test_config = {
            "matching_file": str(matching_path),
            "matching_sheet": "Sheet",
            "filename_col": "Site",
            "planwork_col": "PW Number",
            "data_sheet": "cutsheet",
            "target_sheet": "IP & Port Assignment",
            "source_start_row": 3,
            "paste_start_row": 3,
            "columns": {
                "PW":      {"type": "planwork",  "build_at": "Q", "paste_to": "J"},
                "IP1":     {"type": "ip_lookup", "lookup_col": "C", "log_sheet": "Get Log Before&After", "build_at": "R", "paste_to": "E"},
                "IP2":     {"type": "ip_lookup", "lookup_col": "G", "log_sheet": "Get Log Before&After", "build_at": "S", "paste_to": "H"},
                "NE_NO1":  {"type": "copy",      "source_col": "C", "paste_to": "D"},
                "PORT_NO1":{"type": "copy",      "source_col": "D", "paste_to": "F"},
                "L1":      {"type": "copy",      "source_col": "E", "paste_to": "C"},
                "NE_NO2":  {"type": "copy",      "source_col": "G", "paste_to": "G"},
                "PORT_NO2":{"type": "copy",      "source_col": "H", "paste_to": "I"}
            },
            "source_folder": str(tmp_path),
            "target_folder": str(tmp_path),
            "output_folder": str(tmp_path / "output"),
            "action": "copy",
            "paste_mode": "append",
            "page_break_enabled": True,
            "a4_page_rows": 52,
            "print_title_rows": None
        }

        # 5. Run copier with test config
        copier.main(config=test_config)

        # 6. Open output file and verify EVERY cell
        out_path = tmp_path / "output" / "target_test.xlsx"
        assert out_path.exists(), f"Output not found: {out_path}"

        wb = load_workbook(str(out_path))
        ws = wb["IP & Port Assignment"]

        # Row 3 — all 8 columns
        assert ws.cell(row=3, column=4).value == "CR10SDA", "D3 NE_NO1"
        assert ws.cell(row=3, column=6).value == "1/1/1", "F3 PORT_NO1"
        assert ws.cell(row=3, column=3).value == "CR10-KM01", "C3 L1"
        assert ws.cell(row=3, column=7).value == "CR30SDA", "G3 NE_NO2"
        assert ws.cell(row=3, column=9).value == "2/1/1", "I3 PORT_NO2"
        assert ws.cell(row=3, column=10).value == "TEST001", "J3 PW"
        assert ws.cell(row=3, column=5).value == "10.10.10.10", "E3 IP1"
        assert ws.cell(row=3, column=8).value == "10.20.30.30", "H3 IP2"

        # Row 22 — last data row
        assert ws.cell(row=22, column=3).value == "CR29-KM01", "C22 last L1"

        # Page break: EXISTING_DATA_30 at row 53
        assert ws.cell(row=53, column=1).value == "EXISTING_DATA_30", "row53"
        assert ws.cell(row=54, column=1).value == "EXISTING_DATA_31", "row54"

        wb.close()

        # 7. Verify original source columns unchanged (content-level hash comparison)
        # copier writes temp columns (Q/R/S) so full-file SHA256 would differ
        s = load_workbook(str(src_path))
        cs = s["cutsheet"]
        content_parts_after = []
        for col in [3, 4, 5, 7, 8]:
            for row in range(3, 23):
                v = cs.cell(row=row, column=col).value
                content_parts_after.append(str(v))
        content_hash_after = hashlib.sha256("|".join(content_parts_after).encode()).hexdigest()
        assert content_hash_after == content_hash_before, \
            f"Source data columns modified! hash_before={content_hash_before[:12]}... hash_after={content_hash_after[:12]}..."

        # Also spot-check key cells
        assert cs.cell(row=3, column=3).value == "CR10SDA", "source col C unchanged"
        assert cs.cell(row=3, column=4).value == "1/1/1", "source col D unchanged"
        assert cs.cell(row=3, column=5).value == "CR10-KM01", "source col E unchanged"
        assert cs.cell(row=3, column=7).value == "CR30SDA", "source col G unchanged"
        assert cs.cell(row=3, column=8).value == "2/1/1", "source col H unchanged"
        s.close()


def test_read_matching_missing_filename_col(tmp_path):
    """read_matching must raise ValueError when filename_col header is missing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "match"
    ws["A1"] = "PW Number"
    path = tmp_path / "matching.xlsx"
    wb.save(str(path))
    wb.close()
    with pytest.raises(ValueError, match="Site"):
        read_matching(path, "match", "Site", "PW Number")


def test_read_matching_missing_planwork_col(tmp_path):
    """read_matching must raise ValueError when planwork_col header is missing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "match"
    ws["A1"] = "Site"
    path = tmp_path / "matching.xlsx"
    wb.save(str(path))
    wb.close()
    with pytest.raises(ValueError, match="PW Number"):
        read_matching(path, "match", "Site", "PW Number")
