from pathlib import Path
from openpyxl import load_workbook


def process_workbook(xlsx_path: Path, output_path: Path, replacements: dict[str, str], match_mode: str = "first") -> int:
    """Edit cells and save to output path. Returns number of cells changed.
    
    For each cell matching a prefix, replaces the cell to its RIGHT
    (skipping past merged cells). 
    match_mode: 'first' = one match per prefix, 'all' = every occurrence.
    """
    wb = load_workbook(xlsx_path)
    changed = 0
    matched = set()  # track prefixes already matched (for 'first' mode)
    
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value)
                for prefix, replacement in replacements.items():
                    if match_mode == "first" and prefix in matched:
                        continue
                    if text.startswith(prefix):
                        # Find first non-merged cell to the right
                        right_col = cell.column + 1
                        for merged in ws.merged_cells.ranges:
                            if (merged.min_row <= cell.row <= merged.max_row and
                                merged.min_col <= right_col <= merged.max_col):
                                right_col = merged.max_col + 1
                        ws.cell(row=cell.row, column=right_col).value = replacement
                        changed += 1
                        if match_mode == "first":
                            matched.add(prefix)
                        break
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    wb.close()
    return changed
