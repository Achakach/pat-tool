import io
import random
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from PIL import Image as PILImage


def make_png_rgb(r, g, b, width=40, height=30):
    """Create a PNG with random noise so the file exceeds 500 bytes.

    Solid-color PNGs compress to ~160 bytes regardless of dimensions.
    Adding per-pixel noise defeats PNG compression and pushes the file
    size well over 500 bytes, allowing the noise filter to pass them.
    """
    img = PILImage.new("RGB", (width, height))
    pixels = img.load()
    for x in range(width):
        for y in range(height):
            nr = min(255, max(0, r + random.randint(-40, 40)))
            ng = min(255, max(0, g + random.randint(-40, 40)))
            nb = min(255, max(0, b + random.randint(-40, 40)))
            pixels[x, y] = (nr, ng, nb)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def make_image(buf, col, row, col_off=0, row_off=0):
    """Create openpyxl Image with TwoCellAnchor at given col/row (0-indexed)."""
    xl_img = XLImage(buf)
    marker = AnchorMarker(col=col, colOff=col_off, row=row, rowOff=row_off)
    xl_img.anchor = TwoCellAnchor(_from=marker)
    return xl_img


wb = Workbook()

# ── Sheet "Sales" ──
ws_sales = wb.active
ws_sales.title = "Sales"
ws_sales["A1"] = "Revenue Chart"
ws_sales["E1"] = "Growth Trend"
ws_sales.add_image(make_image(make_png_rgb(70, 130, 180), col=0, row=1))
ws_sales.add_image(make_image(make_png_rgb(34, 139, 34), col=4, row=1))

# ── Sheet "Empty" ──
ws_empty = wb.create_sheet("Empty")
ws_empty.add_image(make_image(make_png_rgb(128, 128, 128), col=1, row=4))

# ── Sheet "Edge" ──
ws_edge = wb.create_sheet("Edge")
ws_edge.add_image(make_image(make_png_rgb(200, 50, 50), col=0, row=0))
ws_edge["B10"] = "Deep Label"
ws_edge.add_image(make_image(make_png_rgb(255, 165, 0), col=1, row=10))

out_path = r"C:\Users\kacha\OneDrive\Desktop\PAT tool\test_fixture.xlsx"
wb.save(out_path)
print(f"Created: {out_path}")
print("Sheets:", wb.sheetnames)
print("Sales A1:", ws_sales["A1"].value)
print("Sales E1:", ws_sales["E1"].value)
print("Empty B4:", ws_empty["B4"].value)
print("Edge B10:", ws_edge["B10"].value)
print("Edge A1:", ws_edge["A1"].value)
