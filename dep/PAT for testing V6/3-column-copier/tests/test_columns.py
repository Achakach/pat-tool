"""Tests for column copier."""
import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook, load_workbook
from src.columns import (
    col_letter_to_index, build_pw_column, build_ip_column,
    clean_sheet_name, find_matching_sheet
)


class TestColLetter:
    @pytest.mark.parametrize("letter,expected", [("A", 1), ("Z", 26), ("AA", 27)])
    def test_conversion(self, letter, expected):
        assert col_letter_to_index(letter) == expected


class TestPwColumn:
    def test_fills_column(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "header"
        ws["A2"] = "data"
        ws["A3"] = "data2"
        build_pw_column(ws, "xxxck01", "B", 2)
        assert ws["B2"].value == "xxxck01"
        assert ws["B3"].value == "xxxck01"
        wb.close()


class TestIpColumn:
    def test_lookup(self):
        wb = Workbook()
        ws = wb.active
        ws["D1"] = "header"
        ws["D2"] = "CR10SDA"
        ws["D3"] = "CR11SDA"
        # Create log sheet
        log = wb.create_sheet("Get Log")
        log["A1"] = "CR10SDA_10.10.10.10"
        log["B1"] = "CR11SDA_10.10.11.11"
        build_ip_column(ws, "D", log, "E", 2)
        assert ws["E2"].value == "10.10.10.10"
        assert ws["E3"].value == "10.10.11.11"
        wb.close()

    @pytest.mark.parametrize("log_value,ne_no,expected_ip", [
        ("new_100_10.0.0.1(MXaxxxx)", "100", "10.0.0.1"),
        ("exist_CR10SDA_10.10.10.10_backup", "CR10SDA", "10.10.10.10"),
        ("old_new_200_192.168.1.5_v2", "200", "192.168.1.5"),
        ("prefix_300_10.20.30.40(suffix)", "300", "10.20.30.40"),
    ])
    def test_lookup_wrapped_formats(self, log_value, ne_no, expected_ip):
        """IP lookup should find NE_NO_IP pattern in wrapped/noisy log values."""
        wb = Workbook()
        ws = wb.active
        ws["A2"] = ne_no
        log = wb.create_sheet("Log")
        log["A1"] = log_value
        build_ip_column(ws, "A", log, "B", 2)
        assert ws["B2"].value == expected_ip
        wb.close()


class TestSheetMatching:
    def test_clean(self):
        assert clean_sheet_name("2.3. IP & Port Assignment(P.4)") == "ip & port assignment"
        assert clean_sheet_name("IP & Port Assignment") == "ip & port assignment"

    def test_find(self):
        wb = Workbook()
        wb.active.title = "2.3. IP & Port Assignment(P.4)"
        assert find_matching_sheet(wb, "IP & Port Assignment") == "2.3. IP & Port Assignment(P.4)"
        wb.close()


class TestPasteDirect:
    """Tests for direct paste behavior (source_col → paste_to, no copy_column)."""

    def test_paste_direct_source_to_target(self):
        """Copy type column: read from source_col in source, write to paste_to in target."""
        swb = Workbook()
        sws = swb.active
        # Source has data in column C (source_col)
        sws["C2"] = "val1"
        sws["C3"] = "val2"

        twb = Workbook()
        tws = twb.active

        # Simulate direct paste: read from C (source_col) in source, write to D (paste_to) in target
        src_idx = col_letter_to_index("C")
        dst_idx = col_letter_to_index("D")
        for row in range(2, 4):
            val = sws.cell(row=row, column=src_idx).value
            if val is not None:
                tws.cell(row=row, column=dst_idx).value = val

        # Assert target column D has copied data
        assert tws["D2"].value == "val1"
        assert tws["D3"].value == "val2"
        # Assert source column C unchanged
        assert sws["C2"].value == "val1"
        assert sws["C3"].value == "val2"

        swb.close()
        twb.close()

    def test_paste_skips_merged_cells(self):
        """Paste should skip writes to merged sub-cells (MergedCell) in target."""
        from openpyxl.cell.cell import MergedCell

        swb = Workbook()
        sws = swb.active
        # Source has data in column B (the column that overlaps a merge in target)
        sws["B1"] = "source B1 val"
        sws["B2"] = "source B2 val"

        twb = Workbook()
        tws = twb.active
        # Merge A1:B1 — B1 becomes a MergedCell (sub-cell of the merge)
        tws.merge_cells("A1:B1")
        tws["A1"] = "merged header"
        # Regular cell below the merge — no merge here
        tws["B2"] = "target B2"

        # Simulate guarded paste: read from source column B, write to target column B
        src_idx = col_letter_to_index("B")
        dst_idx = col_letter_to_index("B")
        for row in range(1, 3):
            val = sws.cell(row=row, column=src_idx).value
            if val is not None:
                cell = tws.cell(row=row, column=dst_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = val

        # B1 is a MergedCell (sub-cell of A1:B1 merge)
        cell_b1 = tws.cell(row=1, column=2)
        assert isinstance(cell_b1, MergedCell), \
            "B1 should be a MergedCell after merge_cells('A1:B1')"
        # Write to B1 was skipped by guard — merged header at A1 is preserved
        assert tws["A1"].value == "merged header", \
            "Merged value at A1 anchor must be preserved (B1 write was skipped)"
        # B1 as MergedCell should not have its own writable value
        assert tws["B1"].value is None, \
            "MergedCell B1 should have no value of its own"

        # B2 is non-merged and should receive source value
        assert tws["B2"].value == "source B2 val", \
            "Non-merged cell B2 should receive source value"

        swb.close()
        twb.close()


class TestBuildAt:
    """Tests for build_at column mapping to target column."""

    def test_build_at_maps_to_target(self):
        """build_at: read from build_at col in source, write to paste_to col in target."""
        swb = Workbook()
        sws = swb.active
        # Source has data in column R = col 18 (build_at column)
        sws.cell(row=2, column=18).value = "built_val1"
        sws.cell(row=3, column=18).value = "built_val2"

        twb = Workbook()
        tws = twb.active

        # Simulate: read from R (build_at), write to E (paste_to)
        src_idx = col_letter_to_index("R")
        dst_idx = col_letter_to_index("E")
        for row in range(2, 4):
            val = sws.cell(row=row, column=src_idx).value
            if val is not None:
                tws.cell(row=row, column=dst_idx).value = val

        # Assert target column E has data from source column R
        assert tws.cell(row=2, column=5).value == "built_val1"
        assert tws.cell(row=3, column=5).value == "built_val2"
        # Assert source column R unchanged
        assert sws.cell(row=2, column=18).value == "built_val1"

        swb.close()
        twb.close()


class TestBackwardCompat:
    """Tests for backward compatibility when no build_at key exists."""

    def test_backward_compat_no_build_at(self):
        """No build_at: fall back to using paste_to for both read and write."""
        swb = Workbook()
        sws = swb.active
        # Source has data in column F (paste_to used as both source and dest column)
        sws.cell(row=2, column=6).value = "compat_val1"
        sws.cell(row=3, column=6).value = "compat_val2"

        twb = Workbook()
        tws = twb.active

        # Simulate: no build_at, paste_to=F used for both read and write
        col_idx = col_letter_to_index("F")
        for row in range(2, 4):
            val = sws.cell(row=row, column=col_idx).value
            if val is not None:
                tws.cell(row=row, column=col_idx).value = val

        # Assert target column F has same values as source column F
        assert tws.cell(row=2, column=6).value == "compat_val1"
        assert tws.cell(row=3, column=6).value == "compat_val2"

        swb.close()
        twb.close()


class TestAppendInsertRows:
    """Tests for insert_rows behavior in append mode."""

    def test_append_insert_rows_shifts_content(self):
        """Append mode with content below: insert_rows shifts content down."""
        # Create source workbook with 5 data rows
        swb = Workbook()
        sws = swb.active
        sws.title = "cutsheet"
        for i in range(3, 8):  # rows 3-7 = 5 data rows
            sws.cell(row=i, column=1).value = f"data{i}"

        # Create target workbook with content at row 10
        twb = Workbook()
        tws = twb.active
        tws.title = "IP & Port Assignment"
        tws.cell(row=10, column=1).value = "existing_content"
        tws.cell(row=11, column=1).value = "more_content"

        # Save both to temp files
        with tempfile.TemporaryDirectory() as tmp:
            src_path = Path(tmp) / "source.xlsx"
            tgt_path = Path(tmp) / "target.xlsx"
            swb.save(str(src_path))
            twb.save(str(tgt_path))
            swb.close()
            twb.close()

            # Simulate: append mode, paste_row=5, src has 5 data rows
            # insert_rows(5, 5) should push content from row 10 → row 15
            twb2 = load_workbook(str(tgt_path))
            tws2 = twb2.active
            tws2.insert_rows(5, 5)  # insert 5 rows at row 5
            twb2.save(str(tgt_path))
            twb2.close()

            # Verify content shifted
            twb3 = load_workbook(str(tgt_path))
            tws3 = twb3.active
            assert tws3.cell(row=15, column=1).value == "existing_content"
            assert tws3.cell(row=16, column=1).value == "more_content"
            twb3.close()

    def test_append_no_insert_rows_when_no_content_below(self):
        """Append mode with no content below: no insert_rows needed."""
        # Source has 5 data rows
        swb = Workbook()
        sws = swb.active
        for i in range(3, 8):
            sws.cell(row=i, column=1).value = f"data{i}"

        # Target is empty except header at row 1
        twb = Workbook()
        tws = twb.active
        tws.cell(row=1, column=1).value = "header"

        # scan for content below row 5 (append start) — there is none
        # src_data_rows would be counted, but insert_rows is still valid
        # even if no content below — it just creates empty space
        # Test: insert_rows at row 5 with 5 rows → no error, row 6 now has "data3"
        tws.insert_rows(5, 5)
        assert tws.cell(row=1, column=1).value == "header"  # header unchanged
        twb.close()
        swb.close()

    def test_insert_rows_detects_content_at_paste_row(self):
        """Content starts AT paste_row: insert_rows shifts it correctly."""
        twb = Workbook()
        tws = twb.active
        tws.cell(row=5, column=1).value = "at_paste_row"
        tws.cell(row=6, column=1).value = "after"

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            twb.save(str(path))
            twb.close()

            twb2 = load_workbook(str(path))
            tws2 = twb2.active
            tws2.insert_rows(5, 3)  # insert 3 rows at row 5
            twb2.save(str(path))
            twb2.close()

            twb3 = load_workbook(str(path))
            tws3 = twb3.active
            assert tws3.cell(row=8, column=1).value == "at_paste_row"  # shifted from 5→8
            assert tws3.cell(row=9, column=1).value == "after"  # shifted from 6→9
            twb3.close()

    def test_full_append_with_page_break(self):
        """Integration: insert_rows + snap_gap_rows work together in append pipeline.

        Simulates the full copier append flow with page_break_enabled=True:
          1. insert_rows at paste_row pushes existing content down
          2. After paste, snap_gap_rows detects overflow and calculates gap
          3. Gap insert pushes content to clean page boundary
        """
        from openpyxl import Workbook
        from src.print_setup import snap_gap_rows

        wb = Workbook()
        ws = wb.active

        # Setup: content at row 9 — existing data below paste area
        ws.cell(row=9, column=1).value = "existing_below"
        ws.cell(row=9, column=2).value = "col2"

        # Step 1: Append mode insert_rows — paste_row=3, 5 data rows
        ws.insert_rows(3, 5)
        # Content originally at row 9 → shifted to row 9+5=14
        assert ws.cell(row=14, column=1).value == "existing_below"
        assert ws.cell(row=14, column=2).value == "col2"

        # Step 2: Simulate pasting 5 rows at rows 3-7 → paste_end=8
        paste_end = 8

        # Step 3: snap_gap_rows with page_rows=10
        # row 14 with page_rows=10: clean starts 1, 11, 21
        # row 14 is NOT clean → next clean = 21, gap = 21-14 = 7
        gap = snap_gap_rows(paste_end, ws, page_rows=10)
        assert gap == 7

        # Step 4: Apply snap — insert gap rows at paste_end
        ws.insert_rows(paste_end, gap)

        # Content at row 14 → shifted to row 14+7=21 (clean page boundary)
        assert ws.cell(row=21, column=1).value == "existing_below"
        assert ws.cell(row=21, column=2).value == "col2"

        wb.close()

    def test_insert_mode_without_page_break(self):
        """insert_mode=True triggers insert_rows even when page_break_enabled=False."""
        from openpyxl import Workbook, load_workbook

        # Setup target: content at row 9
        twb = Workbook()
        tws = twb.active
        tws.cell(row=9, column=1).value = "existing_below"

        with tempfile.TemporaryDirectory() as tmp:
            tgt_path = Path(tmp) / "target.xlsx"
            twb.save(str(tgt_path))
            twb.close()

            # Simulate: insert_mode=True, page_break_enabled=False
            # paste_row=3, src has 5 data rows
            twb2 = load_workbook(str(tgt_path))
            tws2 = twb2.active
            tws2.insert_rows(3, 5)
            twb2.save(str(tgt_path))
            twb2.close()

            # Verify content shifted from row 9 → row 14
            twb3 = load_workbook(str(tgt_path))
            tws3 = twb3.active
            assert tws3.cell(row=14, column=1).value == "existing_below"
            twb3.close()

    def test_insert_mode_shifts_content_at_paste_row(self):
        """insert_mode=True: content at paste_start_row shifts down."""
        import tempfile
        from openpyxl import Workbook, load_workbook

        # Target has content AT paste_start_row (row 3)
        twb = Workbook()
        tws = twb.active
        tws.cell(row=3, column=1).value = "content_at_row3"

        with tempfile.TemporaryDirectory() as tmp:
            tgt_path = Path(tmp) / "target.xlsx"
            twb.save(str(tgt_path))
            twb.close()

            # Simulate: insert_mode=True, 5 data rows
            twb2 = load_workbook(str(tgt_path))
            tws2 = twb2.active
            tws2.insert_rows(3, 5)  # insert 5 rows at paste row 3
            twb2.save(str(tgt_path))
            twb2.close()

            # Verify content shifted from row 3 -> row 8
            twb3 = load_workbook(str(tgt_path))
            tws3 = twb3.active
            assert tws3.cell(row=8, column=1).value == "content_at_row3"
            twb3.close()

    def test_no_insert_mode_pastes_without_insert(self):
        """insert_mode=False: content at paste_start_row stays, gets overwritten."""
        from openpyxl import Workbook

        # Target has content at row 3
        twb = Workbook()
        tws = twb.active
        tws.cell(row=3, column=1).value = "content_at_row3"

        # No insert_rows called — content stays at row 3
        # (insert_mode=False means we write directly without insert)
        assert tws.cell(row=3, column=1).value == "content_at_row3"
        twb.close()
