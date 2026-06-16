"""Tests for print_setup module — A4 page setup and header parsing."""
import pytest
from openpyxl import Workbook
from src.print_setup import (
    _setup_a4_print, _calc_page_rows,
    _parse_print_title_rows, snap_gap_rows
)


class TestSetupA4Print:
    def test_sets_properties(self):
        """_setup_a4_print(ws) sets A4 portrait, auto page breaks, 0.5 top margin."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws)
        assert ws.page_setup.paperSize == 9
        assert ws.page_setup.orientation == 'portrait'
        assert ws.page_setup.autoPageBreaks is True
        assert ws.page_margins.top == 0.5
        wb.close()

    def test_sets_print_title_rows(self):
        """_setup_a4_print(ws, '1:3') → print_title_rows == '$1:$3'."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws, "1:3")
        # openpyxl stores with $ signs
        assert ws.print_title_rows == "$1:$3"
        wb.close()

    def test_none_does_not_set(self):
        """_setup_a4_print(ws, None) → print_title_rows remains None."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws, None)
        assert ws.print_title_rows is None
        wb.close()


class TestCalcPageRows:
    def test_returns_52(self):
        """_calc_page_rows returns ~52 for A4 with 0.5" margins."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws)
        result = _calc_page_rows(ws)
        assert isinstance(result, int)
        # A4 841.89pts, 1" margins=72pts, printable=769.89, /15=51.33→ceil=52
        assert 49 <= result <= 53
        wb.close()


class TestParsePrintTitleRows:
    def test_valid_1_to_6(self):
        """_parse_print_title_rows('1:6') → (6, '1:6')."""
        result = _parse_print_title_rows("1:6")
        # Fix #8: header_count = end-start+1 = 6-1+1 = 6
        assert result == (6, "1:6")

    def test_valid_3_to_8(self):
        """_parse_print_title_rows('3:8') → (6, '3:8')."""
        result = _parse_print_title_rows("3:8")
        # Fix #8: header_count = 8-3+1 = 6, NOT 8
        assert result == (6, "3:8")

    def test_guard_too_many_headers(self, capsys):
        """_parse_print_title_rows('1:52', page_rows=52) → (0, None) with warning."""
        result = _parse_print_title_rows("1:52", page_rows=52)
        assert result == (0, None)
        captured = capsys.readouterr()
        assert "WARNING" in captured.err
