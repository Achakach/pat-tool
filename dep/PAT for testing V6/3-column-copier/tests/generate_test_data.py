#!/usr/bin/env python3
"""Generate test data for copier E2E pipeline."""
import argparse
import sys
from pathlib import Path
from openpyxl import Workbook


def generate_source(output_dir: Path):
    """Create source .xlsx with PW sheet, cutsheet, and log sheet."""
    wb = Workbook()
    # PW sheet
    pw_ws = wb.active
    pw_ws.title = "PW TEST001"
    pw_ws["A1"] = "PW"
    pw_ws["A2"] = "TEST001"

    # cutsheet — 20 rows (rows 3-22)
    cs = wb.create_sheet("cutsheet")
    for i in range(20):
        row = i + 3
        cs.cell(row=row, column=3).value = f"CR{10+i}SDA"       # Col C: NE_NO
        cs.cell(row=row, column=4).value = f"1/1/{i+1}"         # Col D: PORT_NO
        cs.cell(row=row, column=5).value = f"CR{10+i}-KM01"     # Col E: L1
        cs.cell(row=row, column=7).value = f"CR{30+i}SDA"       # Col G: NE_NO2
        cs.cell(row=row, column=8).value = f"2/1/{i+1}"         # Col H: PORT_NO2

    # Get Log Before&After — IP mappings in row 1
    log = wb.create_sheet("Get Log Before&After")
    for i in range(20):
        log.cell(row=1, column=i*2+1).value = f"CR{10+i}SDA_10.10.{10+i}.{10+i}"   # IP1 mappings
        log.cell(row=1, column=i*2+2).value = f"CR{30+i}SDA_10.20.{30+i}.{30+i}"   # IP2 mappings

    path = output_dir / "source_test.xlsx"
    wb.save(str(path))
    wb.close()
    return path


def generate_target(output_dir: Path):
    """Create target .xlsx with IP & Port Assignment sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "IP & Port Assignment"
    ws["A1"] = "Site Info"
    ws.cell(row=30, column=1).value = "EXISTING_DATA_30"
    ws.cell(row=31, column=1).value = "EXISTING_DATA_31"
    # NO merged cells

    path = output_dir / "target_test.xlsx"
    wb.save(str(path))
    wb.close()
    return path


def verify_files(source_path, target_path):
    """Self-verify generated files have correct data."""
    from openpyxl import load_workbook
    s = load_workbook(str(source_path))
    assert "PW TEST001" in s.sheetnames
    assert "cutsheet" in s.sheetnames
    assert "Get Log Before&After" in s.sheetnames
    cs = s["cutsheet"]
    assert cs.cell(row=3, column=3).value == "CR10SDA"
    assert cs.cell(row=3, column=5).value == "CR10-KM01"
    assert cs.cell(row=3, column=7).value == "CR30SDA"
    assert cs.cell(row=22, column=8).value == "2/1/20"
    log = s["Get Log Before&After"]
    assert log.cell(row=1, column=1).value == "CR10SDA_10.10.10.10"
    assert log.cell(row=1, column=2).value == "CR30SDA_10.20.30.30"
    log_max_col = log.max_column
    assert log_max_col >= 40, f"Expected >=40 IP mappings, got {log_max_col}"
    s.close()

    t = load_workbook(str(target_path))
    assert "IP & Port Assignment" in t.sheetnames
    ts = t["IP & Port Assignment"]
    assert ts.cell(row=30, column=1).value == "EXISTING_DATA_30"
    assert ts.cell(row=31, column=1).value == "EXISTING_DATA_31"
    assert len(ts.merged_cells.ranges) == 0, "Expected 0 merged cells"
    t.close()
    print("VERIFY OK")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--tmp-dir", default="tests/tmp")
    args = parser.parse_args()

    output_dir = Path(args.tmp_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Generating test data in {output_dir}")
    src_path = generate_source(output_dir)
    print(f"  Source: {src_path}")
    tgt_path = generate_target(output_dir)
    print(f"  Target: {tgt_path}")

    print("Verifying...")
    verify_files(src_path, tgt_path)
    print("Done.")


if __name__ == "__main__":
    main()
