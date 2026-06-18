"""E2E pipeline test — exercises all 5 tools via run.py with realistic fixtures.

Creates all fixtures programmatically, writes pipeline.json, and runs run.py
as a subprocess. Verifies each stage's outputs exist and the pipeline exits 0.
"""

import json
import sys
import subprocess
import shutil
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlImage

# ── Minimal 1×1 pixel PNG ───────────────────────────────────────────────
PNG_DATA = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
    b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00"
    b"\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82"
)

ROOT = Path(__file__).parent.parent


# ── helpers ──────────────────────────────────────────────────────────────

def _posix(p):
    """Convert Path to forward-slash string for safe embedding in Python source."""
    return p.as_posix()


def _write_config_json(path, config_dict):
    """Write a config dict as JSON file (forward-slashed paths)."""
    path.write_text(json.dumps(config_dict, indent=2), encoding="utf-8")


def _write_stage_script(path, tool_dir, module, config_json_path):
    """Write a Python script that loads config from JSON and calls main()."""
    code = "import sys, json\n"
    code += (
        f'sys.path.insert(0, "{_posix(tool_dir)}")\n'
        f"from {module} import main\n"
        f'with open("{_posix(config_json_path)}") as f:\n'
        "    config = json.load(f)\n"
        "main(config=config)\n"
    )
    path.write_text(code, encoding="utf-8")


# ── E2E test ─────────────────────────────────────────────────────────────

def test_full_pipeline_e2e(tmp_path):
    """Complete 5-stage pipeline runs end-to-end with all real tools."""

    # ── Tool source directories (real project) ──────────────────────
    T1 = ROOT / "1-png-extractor"
    T2 = ROOT / "2-template-generator"
    T3 = ROOT / "3-column-copier"
    T4 = ROOT / "4-cell-editor"
    T5 = ROOT / "5-png-inserter"

    # ── Fixture directories ─────────────────────────────────────────
    t1_in = tmp_path / "t1_input"
    t1_out = tmp_path / "t1_output"
    t2_out = tmp_path / "t2_output"
    t3_src = tmp_path / "t3_source"
    t3_tgt = tmp_path / "t3_target"
    t3_out = tmp_path / "t3_output"
    t4_in = tmp_path / "t4_input"
    t4_out = tmp_path / "t4_output"
    t5_png = tmp_path / "t5_png"
    t5_xlsx = tmp_path / "t5_xlsx"
    t5_out = tmp_path / "t5_output"

    for d in [t1_in, t1_out, t2_out, t3_src, t3_tgt, t3_out,
              t4_in, t4_out, t5_png, t5_xlsx, t5_out]:
        d.mkdir(parents=True, exist_ok=True)

    # ── Write PNG test file ─────────────────────────────────────────
    png_file = tmp_path / "_test.png"
    png_file.write_bytes(PNG_DATA)

    # ═══════════════════════════════════════════════════════════════
    # Fixture 1: matching.xlsx
    # ═══════════════════════════════════════════════════════════════
    matching = tmp_path / "matching.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Site"
    ws["B1"] = "PW Number"
    ws["A2"] = "Alpha"
    ws["B2"] = "XX001"
    wb.save(str(matching))
    wb.close()

    # ═══════════════════════════════════════════════════════════════
    # Fixture 2: template.xlsx
    # ═══════════════════════════════════════════════════════════════
    template = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "PortAssignment"         # sheet name fuzzy-matched by tools 3,5
    ws["A1"] = "name:"                  # prefix for tool 4 replacement
    ws["A3"] = "Data"
    wb.save(str(template))
    wb.close()

    # ═══════════════════════════════════════════════════════════════
    # Fixture 3: Source XLSX (for tools 1 + 3)
    # ═══════════════════════════════════════════════════════════════
    source_xlsx = t1_in / "cutsheet_source.xlsx"
    swb = Workbook()

    # Sheet "exist Alpha" — with label text + embedded image
    ws1 = swb.active
    ws1.title = "exist Alpha"
    ws1["A1"] = "Header"
    ws1["A3"] = "PortAssignment"        # label text: tool 1 naming → tool 5 sheet match
    img = XlImage(str(png_file))
    ws1.add_image(img, "B4")            # anchor at B4 (0-idx row=3 col=1)

    # Sheet "PW XX001" — planwork sheet
    ws2 = swb.create_sheet("PW XX001")
    ws2["A1"] = "Planwork Sheet"

    # Sheet "Cutsheet" — data for tool 3 column copying
    ws3 = swb.create_sheet("Cutsheet")
    ws3["C1"] = "NE_NO"
    ws3["D1"] = "PORT_NO"
    ws3["E1"] = "L1_NAME"
    ws3["G1"] = "NE_NO2"
    ws3["H1"] = "PORT_NO2"
    # Data row at row 3 (source_start_row=3)
    ws3["C3"] = "NE001"
    ws3["D3"] = "P001"
    ws3["E3"] = "L1_Alpha"
    ws3["G3"] = "NE002"
    ws3["H3"] = "P002"

    swb.save(str(source_xlsx))
    swb.close()

    # Also put a copy in t3_source (tool 3 reads from there)
    shutil.copy2(str(source_xlsx), str(t3_src / "cutsheet_source.xlsx"))

    # ═══════════════════════════════════════════════════════════════
    # Copy run.py into tmp_path
    # ═══════════════════════════════════════════════════════════════
    run_py = tmp_path / "run.py"
    shutil.copy2(ROOT / "run.py", run_py)

    # ═══════════════════════════════════════════════════════════════
    # Write stage configs (JSON) + scripts (Python)
    # ═══════════════════════════════════════════════════════════════

    # ── Stage 1: PNG Extractor ──────────────────────────────────────
    s1_cfg = tmp_path / "_s1_cfg.json"
    _write_config_json(s1_cfg, {
        "input_folder": _posix(t1_in),
        "output_folder": _posix(t1_out),
        "noise_threshold": 0,
    })
    s1_script = tmp_path / "_s1.py"
    _write_stage_script(s1_script, T1, "extract_pngs", s1_cfg)

    # ── Stage 2: Template Generator ─────────────────────────────────
    s2_cfg = tmp_path / "_s2_cfg.json"
    _write_config_json(s2_cfg, {
        "matching_file": _posix(matching),
        "template": _posix(template),
        "output_folder": _posix(t2_out),
        "matching_sheet": "Sheet1",
        "filename_col": "Site",
    })
    s2_script = tmp_path / "_s2.py"
    _write_stage_script(s2_script, T2, "generate", s2_cfg)

    # ── Stage 3: Column Copier ──────────────────────────────────────
    s3_cfg = tmp_path / "_s3_cfg.json"
    _write_config_json(s3_cfg, {
        "action": "copy",
        "matching_file": _posix(matching),
        "matching_sheet": "Sheet1",
        "filename_col": "Site",
        "planwork_col": "PW Number",
        "data_sheet": "Cutsheet",
        "target_sheet": "PortAssignment",
        "source_start_row": 3,
        "paste_start_row": 3,
        "page_break_enabled": False,
        "a4_page_rows": None,
        "print_title_rows": None,
        "columns": {
            "PW":      {"type": "planwork", "build_at": "Q", "paste_to": "J"},
            "NE_NO1":  {"type": "copy", "source_col": "C", "paste_to": "D"},
            "PORT_NO1":{"type": "copy", "source_col": "D", "paste_to": "F"},
            "L1":      {"type": "copy", "source_col": "E", "paste_to": "C"},
            "NE_NO2":  {"type": "copy", "source_col": "G", "paste_to": "G"},
            "PORT_NO2":{"type": "copy", "source_col": "H", "paste_to": "I"},
        },
        "source_folder": _posix(t3_src),
        "target_folder": _posix(t3_tgt),
        "output_folder": _posix(t3_out),
    })
    s3_script = tmp_path / "_s3.py"
    _write_stage_script(s3_script, T3, "copier", s3_cfg)

    # ── Stage 4: Cell Editor ────────────────────────────────────────
    s4_cfg = tmp_path / "_s4_cfg.json"
    _write_config_json(s4_cfg, {
        "input_folder": _posix(t4_in),
        "output_folder": _posix(t4_out),
        "match_mode": "first",
        "replacements": {"name:": "kacha"},
    })
    s4_script = tmp_path / "_s4.py"
    _write_stage_script(s4_script, T4, "edit", s4_cfg)

    # ── Stage 5: PNG Inserter ───────────────────────────────────────
    s5_cfg = tmp_path / "_s5_cfg.json"
    _write_config_json(s5_cfg, {
        "matching_file": _posix(matching),
        "matching_sheet": "Sheet1",
        "filename_col": "Site",
        "planwork_col": "PW Number",
        "xlsx_folder": _posix(t5_xlsx),
        "png_folder": _posix(t5_png),
        "output_folder": _posix(t5_out),
        "purge_from_row": 10,
        "label_merge_to_col": "K",
        "insert_gap_rows": 1,
        "image_insert_col": "C",
        "image_display_width": 300,
        "page_break_before_label": False,
        "print_title_rows": None,
    })
    s5_script = tmp_path / "_s5.py"
    _write_stage_script(s5_script, T5, "insert", s5_cfg)

    # ═══════════════════════════════════════════════════════════════
    # Build pipeline.json
    # ═══════════════════════════════════════════════════════════════
    pipeline = {
        "pipeline": {
            "1-png-extractor": {
                "command": f"python {_posix(s1_script)}",
                "copy": [{"from": "t1_output/*.png", "to": "t5_png/"}],
            },
            "2-template-generator": {
                "command": f"python {_posix(s2_script)}",
                "copy": [{"from": "t2_output/*.xlsx", "to": "t3_target/"}],
            },
            "3-column-copier": {
                "command": f"python {_posix(s3_script)}",
                "copy": [{"from": "t3_output/*.xlsx", "to": "t4_input/"}],
            },
            "4-cell-editor": {
                "command": f"python {_posix(s4_script)}",
                "copy": [{"from": "t4_output/*.xlsx", "to": "t5_xlsx/"}],
            },
            "5-png-inserter": {
                "command": f"python {_posix(s5_script)}",
                "copy": [],
            },
        }
    }
    pipeline_json = tmp_path / "pipeline.json"
    pipeline_json.write_text(json.dumps(pipeline, indent=2), encoding="utf-8")

    # ═══════════════════════════════════════════════════════════════
    # Run the pipeline
    # ═══════════════════════════════════════════════════════════════
    result = subprocess.run(
        [sys.executable, str(run_py)],
        cwd=str(tmp_path),
        capture_output=True,
        text=True,
        timeout=120,
    )

    # Debug output on failure
    if result.returncode != 0:
        print("── STDOUT ──", file=sys.stderr)
        print(result.stdout, file=sys.stderr)
        print("── STDERR ──", file=sys.stderr)
        print(result.stderr, file=sys.stderr)

    # ═══════════════════════════════════════════════════════════════
    # Verify
    # ═══════════════════════════════════════════════════════════════

    # Stage 1: PNG extracted
    pngs = list(t1_out.glob("*.png"))
    assert len(pngs) >= 1, (
        f"Stage 1 FAIL: Expected >=1 PNG in {t1_out}, got {len(pngs)}"
    )

    # Stage 2: Template copies generated
    templates = list(t2_out.glob("*.xlsx"))
    assert len(templates) >= 1, (
        f"Stage 2 FAIL: Expected >=1 template in {t2_out}, got {len(templates)}"
    )
    assert any(f.name == "Alpha.xlsx" for f in templates), (
        f"Stage 2 FAIL: Alpha.xlsx not found in {[f.name for f in templates]}"
    )

    # Stage 3: Target in output with copied data
    t3_files = list(t3_out.glob("*.xlsx"))
    assert len(t3_files) >= 1, (
        f"Stage 3 FAIL: Expected >=1 file in {t3_out}, got {len(t3_files)}"
    )

    # Stage 4: Edited file in output
    t4_files = list(t4_out.glob("*.xlsx"))
    assert len(t4_files) >= 1, (
        f"Stage 4 FAIL: Expected >=1 file in {t4_out}, got {len(t4_files)}"
    )

    # Stage 5: Final output exists
    finals = list(t5_out.glob("*.xlsx"))
    assert len(finals) >= 1, (
        f"Stage 5 FAIL: Expected >=1 final XLSX in {t5_out}, got {len(finals)}"
    )

    # Pipeline exit code
    assert result.returncode == 0, (
        f"Pipeline FAILED with exit code {result.returncode}"
    )

    # ── Verify content of final output ─────────────────────────────
    final_wb = load_workbook(str(finals[0]))
    found_replacement = False
    has_label = False
    has_port_assignment_sheet = False
    total_images = 0

    for sn in final_wb.sheetnames:
        if sn == "PortAssignment":
            has_port_assignment_sheet = True
        sheet = final_wb[sn]
        total_images += len(sheet._images) if hasattr(sheet, '_images') else 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "kacha":
                    found_replacement = True
                if cell.value == "Alpha":
                    has_label = True

    final_wb.close()

    assert found_replacement, (
        "Tool 4 FAIL: 'name:' → 'kacha' replacement not found in final output"
    )
    assert has_label, (
        "Tool 5 FAIL: label 'Alpha' not found in final output"
    )
    assert total_images >= 1, (
        f"Tool 5 FAIL: Expected >=1 image in final XLSX, got {total_images}"
    )
    assert has_port_assignment_sheet, (
        "Tool 5 FAIL: 'PortAssignment' sheet not found in final output"
    )
