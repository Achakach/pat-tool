"""CLI tests for column copier — subprocess integration."""
import json
import shutil
import subprocess
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook

SCRIPT = Path(__file__).parent.parent / "copier.py"
SRC = Path(__file__).parent.parent / "src"


def _setup(tmp_path, config_dict=None):
    d = tmp_path / "tool3"
    d.mkdir()
    shutil.copy2(SCRIPT, d / "copier.py")
    shutil.copytree(SRC, d / "src")
    if config_dict:
        (d / "config.json").write_text(json.dumps(config_dict), encoding="utf-8")
    return d


class TestCli:
    def test_happy_path(self, tmp_path):
        """End-to-end: source → build temp cols → copy to target → output."""
        d = _setup(tmp_path)

        # --- 1. Create matching.xlsx at parent level ---
        matching_path = tmp_path / "matching.xlsx"
        mwb = Workbook()
        mws = mwb.active
        mws.title = "match"
        mws["A1"] = "Site"
        mws["B1"] = "PW Number"
        mws["A2"] = "Alpha"
        mws["B2"] = "XX001"
        mwb.save(str(matching_path))
        mwb.close()

        # --- 2. Create source test.xlsx ---
        src_dir = d / "source"
        src_dir.mkdir()
        swb = Workbook()
        # PW sheet for planwork extraction: regex r'^PW\s+(.+)' matches
        pw_ws = swb.active
        pw_ws.title = "PW XX001"
        # Data sheet with 2 data rows at rows 3-4
        cs = swb.create_sheet("Cutsheet")
        cs["A1"] = "Header"
        cs["A3"] = "data_row3"
        cs["A4"] = "data_row4"
        swb.save(str(src_dir / "test.xlsx"))
        swb.close()

        # --- 3. Create target Alpha.xlsx ---
        tgt_dir = d / "target"
        tgt_dir.mkdir()
        twb = Workbook()
        tws = twb.active
        tws.title = "IP"  # fuzzy-matches config target_sheet "IP"
        twb.save(str(tgt_dir / "Alpha.xlsx"))
        twb.close()

        # --- 4. Write config ---
        config = {
            "action": "copy",
            "matching_file": "../matching.xlsx",
            "matching_sheet": "match",
            "filename_col": "Site",
            "planwork_col": "PW Number",
            "data_sheet": "Cutsheet",
            "target_sheet": "IP",
            "source_start_row": 3,
            "paste_start_row": 3,
            "insert_mode": False,
            "source_folder": "./source",
            "target_folder": "./target",
            "output_folder": "./output",
            "columns": {
                "PW": {"type": "planwork", "build_at": "Q", "paste_to": "J"}
            },
            "page_break_enabled": False,
            "print_title_rows": None,
            "a4_page_rows": None
        }
        (d / "config.json").write_text(json.dumps(config), encoding="utf-8")

        # --- 5. Run ---
        result = subprocess.run(
            [sys.executable, str(d / "copier.py")],
            cwd=str(d),
            capture_output=True,
            text=True,
        )

        # --- 6. Verify ---
        assert result.returncode == 0, (
            f"Exit {result.returncode}:\nSTDERR: {result.stderr}"
        )
        assert "Processing" in result.stdout
        assert "Copied to" in result.stdout

        out_path = d / "output" / "Alpha.xlsx"
        assert out_path.exists(), f"Output not found: {out_path}"

        # Verify content: column J has planwork value pasted from column Q
        owb = load_workbook(str(out_path))
        ows = owb["IP"]
        assert ows.cell(row=3, column=10).value == "XX001", "J3 missing planwork"
        assert ows.cell(row=4, column=10).value == "XX001", "J4 missing planwork"
        owb.close()

    def test_missing_config(self, tmp_path):
        """No config.json → FileNotFoundError → non-zero exit."""
        d = _setup(tmp_path)

        result = subprocess.run(
            [sys.executable, str(d / "copier.py")],
            cwd=str(d),
            capture_output=True,
            text=True,
        )

        assert result.returncode != 0, (
            f"Expected non-zero exit, got {result.returncode}"
        )

    def test_missing_source_folder(self, tmp_path):
        """Source folder doesn't exist → glob finds nothing → exit 0."""
        d = _setup(tmp_path)

        # matching.xlsx at parent level (used by copier before the loop)
        matching_path = tmp_path / "matching.xlsx"
        mwb = Workbook()
        mws = mwb.active
        mws.title = "match"
        mws["A1"] = "Site"
        mws["B1"] = "PW Number"
        mwb.save(str(matching_path))
        mwb.close()

        config = {
            "action": "copy",
            "matching_file": "../matching.xlsx",
            "matching_sheet": "match",
            "filename_col": "Site",
            "planwork_col": "PW Number",
            "data_sheet": "Cutsheet",
            "target_sheet": "IP",
            "source_start_row": 3,
            "paste_start_row": 3,
            "insert_mode": False,
            "source_folder": "./source",
            "target_folder": "./target",
            "output_folder": "./output",
            "columns": {
                "PW": {"type": "planwork", "build_at": "Q", "paste_to": "J"}
            },
            "page_break_enabled": False,
            "print_title_rows": None,
            "a4_page_rows": None
        }
        (d / "config.json").write_text(json.dumps(config), encoding="utf-8")

        # source/ is intentionally NOT created
        # target/ folder not needed (loop never executes)

        result = subprocess.run(
            [sys.executable, str(d / "copier.py")],
            cwd=str(d),
            capture_output=True,
            text=True,
        )

        # Tool 3 returns exit 0 even when source folder is empty
        assert result.returncode == 0, (
            f"Expected exit 0, got {result.returncode}:\nSTDERR: {result.stderr}"
        )
        assert "Processing" not in result.stdout, (
            "No 'Processing' expected when source folder is empty"
        )
