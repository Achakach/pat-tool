"""CLI integration tests for 4-cell-editor via subprocess."""
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

SCRIPT = Path(__file__).parent.parent / "edit.py"
SRC = Path(__file__).parent.parent / "src"


def _setup(tmp_path, config_dict=None):
    """Copy edit.py + src to temp dir. Optionally create config.json."""
    d = tmp_path / "tool4"
    d.mkdir()
    shutil.copy2(SCRIPT, d / "edit.py")
    shutil.copytree(SRC, d / "src")
    if config_dict is not None:
        (d / "config.json").write_text(json.dumps(config_dict), encoding="utf-8")
    return d


def _run(d):
    """Run edit.py subprocess from directory d."""
    return subprocess.run(
        [sys.executable, "edit.py"],
        cwd=str(d),
        capture_output=True,
        text=True,
    )


class TestCLI:

    def test_happy_path(self, tmp_path):
        """Full CLI run: input xlsx → process → output xlsx with replaced cell."""
        d = _setup(tmp_path, {
            "input_folder": "./input",
            "output_folder": "./output",
            "match_mode": "first",
            "replacements": {"name:": "kacha"},
        })
        input_dir = d / "input"
        input_dir.mkdir()

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name:"
        ws["B1"] = "old_value"
        wb.save(str(input_dir / "test.xlsx"))
        wb.close()

        result = _run(d)

        assert result.returncode == 0, f"stderr: {result.stderr}"
        assert "Done" in result.stdout
        assert "Changed 1 cell" in result.stdout

        output_path = d / "output" / "test.xlsx"
        assert output_path.exists()

        wb2 = load_workbook(output_path)
        ws2 = wb2.active
        assert ws2["A1"].value == "name:"       # untouched
        assert ws2["B1"].value == "kacha"        # replaced
        wb2.close()

    def test_missing_config(self, tmp_path):
        """No config.json → subprocess exits non-zero."""
        d = _setup(tmp_path)  # no config_dict

        result = _run(d)

        assert result.returncode != 0

    def test_missing_input_folder(self, tmp_path):
        """Config present but input/ folder missing → exit 2, stderr message."""
        d = _setup(tmp_path, {
            "input_folder": "./input",
            "output_folder": "./output",
            "match_mode": "first",
            "replacements": {"name:": "kacha"},
        })
        # DON'T create input/ directory

        result = _run(d)

        assert result.returncode == 2
        assert "Input folder not found" in result.stderr
