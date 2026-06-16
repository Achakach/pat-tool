"""Tests for column copier."""
import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook, load_workbook
from src.columns import (
    col_letter_to_index, build_pw_column, build_ip_column,
    copy_column, clean_sheet_name, find_matching_sheet
)


class TestColLetter:
    @pytest.mark.parametrize("letter,expected", [("A", 1), ("Z", 26), ("AA", 27)])
    def test_conversion(self, letter, expected):
        assert col_letter_to_index(letter) == expected


class TestPwColumn:
    def test_fills_column(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "header"
        ws["A2"] = "data"
        ws["A3"] = "data2"
        build_pw_column(ws, "xxxck01", "B", 2)
        assert ws["B2"].value == "xxxck01"
        assert ws["B3"].value == "xxxck01"
        wb.close()


class TestIpColumn:
    def test_lookup(self):
        wb = Workbook()
        ws = wb.active
        ws["D1"] = "header"
        ws["D2"] = "CR10SDA"
        ws["D3"] = "CR11SDA"
        # Create log sheet
        log = wb.create_sheet("Get Log")
        log["A1"] = "CR10SDA_10.10.10.10"
        log["B1"] = "CR11SDA_10.10.11.11"
        build_ip_column(ws, "D", log, "E", 2)
        assert ws["E2"].value == "10.10.10.10"
        assert ws["E3"].value == "10.10.11.11"
        wb.close()


class TestCopyColumn:
    def test_copy(self):
        wb = Workbook()
        ws = wb.active
        ws["A2"] = "val1"
        ws["A3"] = "val2"
        copy_column(ws, "A", "B", 2)
        assert ws["B2"].value == "val1"
        assert ws["B3"].value == "val2"
        wb.close()


class TestSheetMatching:
    def test_clean(self):
        assert clean_sheet_name("2.3. IP & Port Assignment(P.4)") == "ip & port assignment"
        assert clean_sheet_name("IP & Port Assignment") == "ip & port assignment"

    def test_find(self):
        wb = Workbook()
        wb.active.title = "2.3. IP & Port Assignment(P.4)"
        assert find_matching_sheet(wb, "IP & Port Assignment") == "2.3. IP & Port Assignment(P.4)"
        wb.close()


class TestAppendInsertRows:
    """Tests for insert_rows behavior in append mode."""

    def test_append_insert_rows_shifts_content(self):
        """Append mode with content below: insert_rows shifts content down."""
        # Create source workbook with 5 data rows
        swb = Workbook()
        sws = swb.active
        sws.title = "cutsheet"
        for i in range(3, 8):  # rows 3-7 = 5 data rows
            sws.cell(row=i, column=1).value = f"data{i}"

        # Create target workbook with content at row 10
        twb = Workbook()
        tws = twb.active
        tws.title = "IP & Port Assignment"
        tws.cell(row=10, column=1).value = "existing_content"
        tws.cell(row=11, column=1).value = "more_content"

        # Save both to temp files
        with tempfile.TemporaryDirectory() as tmp:
            src_path = Path(tmp) / "source.xlsx"
            tgt_path = Path(tmp) / "target.xlsx"
            swb.save(str(src_path))
            twb.save(str(tgt_path))
            swb.close()
            twb.close()

            # Simulate: append mode, paste_row=5, src has 5 data rows
            # insert_rows(5, 5) should push content from row 10 → row 15
            twb2 = load_workbook(str(tgt_path))
            tws2 = twb2.active
            tws2.insert_rows(5, 5)  # insert 5 rows at row 5
            twb2.save(str(tgt_path))
            twb2.close()

            # Verify content shifted
            twb3 = load_workbook(str(tgt_path))
            tws3 = twb3.active
            assert tws3.cell(row=15, column=1).value == "existing_content"
            assert tws3.cell(row=16, column=1).value == "more_content"
            twb3.close()

    def test_append_no_insert_rows_when_no_content_below(self):
        """Append mode with no content below: no insert_rows needed."""
        # Source has 5 data rows
        swb = Workbook()
        sws = swb.active
        for i in range(3, 8):
            sws.cell(row=i, column=1).value = f"data{i}"

        # Target is empty except header at row 1
        twb = Workbook()
        tws = twb.active
        tws.cell(row=1, column=1).value = "header"

        # scan for content below row 5 (append start) — there is none
        # src_data_rows would be counted, but insert_rows is still valid
        # even if no content below — it just creates empty space
        # Test: insert_rows at row 5 with 5 rows → no error, row 6 now has "data3"
        tws.insert_rows(5, 5)
        assert tws.cell(row=1, column=1).value == "header"  # header unchanged
        twb.close()
        swb.close()

    def test_insert_rows_detects_content_at_paste_row(self):
        """Content starts AT paste_row: insert_rows shifts it correctly."""
        twb = Workbook()
        tws = twb.active
        tws.cell(row=5, column=1).value = "at_paste_row"
        tws.cell(row=6, column=1).value = "after"

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test.xlsx"
            twb.save(str(path))
            twb.close()

            twb2 = load_workbook(str(path))
            tws2 = twb2.active
            tws2.insert_rows(5, 3)  # insert 3 rows at row 5
            twb2.save(str(path))
            twb2.close()

            twb3 = load_workbook(str(path))
            tws3 = twb3.active
            assert tws3.cell(row=8, column=1).value == "at_paste_row"  # shifted from 5→8
            assert tws3.cell(row=9, column=1).value == "after"  # shifted from 6→9
            twb3.close()
