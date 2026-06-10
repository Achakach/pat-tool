#!/usr/bin/env python3
"""Generate XLSX files from template using matching file filenames."""

import sys
import json
import shutil
from pathlib import Path
from openpyxl import load_workbook


def main():
    config_path = Path(__file__).parent / "config.json"
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    matching_file = Path(config["matching_file"])
    template_file = Path(config["template"])
    output_folder = Path(config["output_folder"])

    if not matching_file.is_absolute():
        matching_file = (Path(__file__).parent / matching_file).resolve()
    if not template_file.is_absolute():
        template_file = (Path(__file__).parent / template_file).resolve()
    if not output_folder.is_absolute():
        output_folder = (Path(__file__).parent / output_folder).resolve()

    if not matching_file.exists():
        print(f"Matching file not found: {matching_file}", file=sys.stderr)
        sys.exit(1)
    if not template_file.exists():
        print(f"Template file not found: {template_file}", file=sys.stderr)
        sys.exit(1)

    output_folder.mkdir(parents=True, exist_ok=True)

    # Read matching file to get unique filenames
    wb = load_workbook(str(matching_file), data_only=True)
    ws = wb[config["matching_sheet"]]

    # Find columns by header (row 1)
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column

    fn_col = headers.get(config["filename_col"].lower())
    if fn_col is None:
        print(f"Column '{config['filename_col']}' not found in matching file headers", file=sys.stderr)
        sys.exit(1)

    filenames = []
    current_filename = None

    for row in ws.iter_rows(min_row=2):
        cell = row[fn_col - 1] if len(row) >= fn_col else None
        name = str(cell.value).strip() if cell and cell.value else None
        if name:
            current_filename = name
        if current_filename and current_filename not in filenames:
            filenames.append(current_filename)

    wb.close()

    if not filenames:
        print("No filenames found in matching file", file=sys.stderr)
        sys.exit(1)

    generated = 0
    for name in sorted(filenames):
        fname = name if name.endswith(".xlsx") else f"{name}.xlsx"
        out = output_folder / fname
        shutil.copy2(template_file, out)
        print(f"Generated: {fname}")
        generated += 1

    print(f"\nDone. Generated {generated} file(s) from template.")


if __name__ == "__main__":
    main()
