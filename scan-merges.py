"""scan-merges.py -- Scan XLSX for merged cells and print them."""

import sys
from pathlib import Path
from openpyxl import load_workbook

def main():
    if len(sys.argv) != 2:
        print("Usage: python scan-merges.py <path_to_xlsx>")
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.is_file():
        print(f"Error: File not found: {path}")
        sys.exit(1)

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"Error opening workbook: {e}")
        sys.exit(1)

    total_merged = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ranges = list(ws.merged_cells.ranges)
        print(f"Sheet: {sheet_name}")
        if not ranges:
            print("  No merged cells")
            continue
        for r in ranges:
            total_merged += 1
            print(f"  {r}  (rows {r.min_row}-{r.max_row}, cols {r.min_col}-{r.max_col})")

    if total_merged == 0:
        print("No merged cells found.")
    wb.close()

if __name__ == "__main__":
    main()
