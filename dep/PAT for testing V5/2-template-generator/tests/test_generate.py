"""Tests for 2-template-generator — template copy with matching.xlsx filenames."""
import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
from generate import main


# ── helpers ──────────────────────────────────────────────────────────────

def _make_template(path):
    """Minimal template workbook: one sheet, "Header" in A1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Header"
    wb.save(str(path))
    wb.close()


def _make_matching(path, rows, headers=("Site", "PW Number"), sheet="Sheet1"):
    """Create matching.xlsx with headers in row 1 and data rows from row 2."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    for ri, row_data in enumerate(rows, 2):
        for ci, val in enumerate(row_data, 1):
            # Only write non-None values to avoid creating stub cells
            if val is not None:
                ws.cell(row=ri, column=ci, value=val)
    wb.save(str(path))
    wb.close()


def _cfg(tmp_path, **overrides):
    """Build a config dict with paths rooted under tmp_path."""
    return {
        "matching_file": str(tmp_path / "matching.xlsx"),
        "template": str(tmp_path / "template.xlsx"),
        "output_folder": str(tmp_path / "output"),
        "matching_sheet": "Sheet1",
        "filename_col": "Site",
        **overrides,
    }


# ── tests ────────────────────────────────────────────────────────────────

class TestGenerate:
    """All tests use tmp_path — no real files touched."""

    # 1 ────────────────────────────────────────────────────────────────────
    def test_generates_from_matching(self, tmp_path):
        """3 data rows, 2 unique Site values → 2 output .xlsx files."""
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))
        _make_matching(Path(config["matching_file"]), [
            ("SiteA", "PW001"),
            ("SiteB", "PW002"),
            ("SiteA", "PW003"),
        ])

        main(config=config)

        out = Path(config["output_folder"])
        files = sorted(out.glob("*.xlsx"))
        assert [f.name for f in files] == ["SiteA.xlsx", "SiteB.xlsx"]

        # Each output must be a valid copy (same "Header" content)
        for f in files:
            wb = load_workbook(str(f))
            assert wb.active["A1"].value == "Header"
            wb.close()

    # 2 ────────────────────────────────────────────────────────────────────
    def test_blank_site_inherits(self, tmp_path):
        """Empty Site cell inherits filename from the row above."""
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))
        _make_matching(Path(config["matching_file"]), [
            ("Alpha", "PW001"),
            ("", "PW002"),       # inherits "Alpha"
            ("Beta", "PW003"),
            ("", "PW004"),       # inherits "Beta"
        ])

        main(config=config)

        out = Path(config["output_folder"])
        files = sorted(out.glob("*.xlsx"))
        # Alpha.xlsx, Beta.xlsx — no duplicate for inherited rows
        assert [f.name for f in files] == ["Alpha.xlsx", "Beta.xlsx"]

    # 3 ────────────────────────────────────────────────────────────────────
    def test_appends_xlsx_suffix(self, tmp_path):
        """Filename without .xlsx → output gets suffix appended."""
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))
        _make_matching(Path(config["matching_file"]), [
            ("mysite", "PW001"),
        ])

        main(config=config)

        out = Path(config["output_folder"])
        assert (out / "mysite.xlsx").exists()
        assert not (out / "mysite").exists()

    # 4 ────────────────────────────────────────────────────────────────────
    def test_already_has_xlsx_suffix(self, tmp_path):
        """Filename already ends with .xlsx → no double suffix."""
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))
        _make_matching(Path(config["matching_file"]), [
            ("mysite.xlsx", "PW001"),
        ])

        main(config=config)

        out = Path(config["output_folder"])
        assert (out / "mysite.xlsx").exists()
        assert not (out / "mysite.xlsx.xlsx").exists()

    # 5 ────────────────────────────────────────────────────────────────────
    def test_template_not_found(self, tmp_path):
        """Config points to nonexistent template → SystemExit(1)."""
        config = _cfg(tmp_path)
        _make_matching(Path(config["matching_file"]), [
            ("SiteA", "PW001"),
        ])
        # template.xlsx intentionally never created

        with pytest.raises(SystemExit) as exc:
            main(config=config)
        assert exc.value.code == 1

    # 6 ────────────────────────────────────────────────────────────────────
    def test_matching_not_found(self, tmp_path):
        """Config points to nonexistent matching file → SystemExit(1)."""
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))
        # matching.xlsx intentionally never created

        with pytest.raises(SystemExit) as exc:
            main(config=config)
        assert exc.value.code == 1

    # 7 ────────────────────────────────────────────────────────────────────
    def test_no_filenames_found(self, tmp_path):
        """Matching file valid but no Site column data → SystemExit(1)."""
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))
        # Rows exist but Site column is empty across all data rows
        _make_matching(Path(config["matching_file"]), [
            ("", "PW001"),
            ("", "PW002"),
        ])

        with pytest.raises(SystemExit) as exc:
            main(config=config)
        assert exc.value.code == 1

    # 8 ────────────────────────────────────────────────────────────────────
    def test_empty_site_creates_none_file(self, tmp_path):
        """Site cell explicitly None → known bug: str(None) would produce "None.xlsx".

        HOWEVER: the current code guards with `cell and cell.value` — when
        cell.value is None the condition is falsy so str() is never called.
        Therefore the bug does NOT manifest with the current generate.py.

        If the guard were weakened (e.g. `cell is not None`), then
        str(None).strip() = "None" → "None.xlsx" would be created.
        This test verifies the current safe behaviour (no spurious None file).
        """
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))

        # Build matching where the Site cell value is literally Python None
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Site"
        ws["B1"] = "PW Number"
        ws["A2"] = None       # explicitly None
        ws["B2"] = "PW001"
        wb.save(str(config["matching_file"]))
        wb.close()

        # current_filename stays None → filenames empty → SystemExit(1)
        with pytest.raises(SystemExit) as exc:
            main(config=config)
        assert exc.value.code == 1

        # Confirm no "None.xlsx" was created
        out = Path(config["output_folder"])
        assert not (out / "None.xlsx").exists()

    # 9 ────────────────────────────────────────────────────────────────────
    def test_skip_duplicate_filenames(self, tmp_path):
        """Same Site value in 3 rows → only 1 output file."""
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))
        _make_matching(Path(config["matching_file"]), [
            ("DupSite", "PW001"),
            ("DupSite", "PW002"),
            ("DupSite", "PW003"),
        ])

        main(config=config)

        out = Path(config["output_folder"])
        files = list(out.glob("*.xlsx"))
        assert len(files) == 1
        assert files[0].name == "DupSite.xlsx"

    # 10 ───────────────────────────────────────────────────────────────────
    def test_generate_with_config_dict(self, tmp_path):
        """Full integration: 5 rows with inheritance → 3 unique output files."""
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))
        _make_matching(Path(config["matching_file"]), [
            ("Alpha", "PW001"),
            ("Beta", "PW002"),
            ("", "PW003"),       # inherits Beta
            ("Gamma", "PW004"),
            ("", "PW005"),       # inherits Gamma
        ])

        main(config=config)

        out = Path(config["output_folder"])
        files = sorted(out.glob("*.xlsx"))
        assert [f.name for f in files] == [
            "Alpha.xlsx",
            "Beta.xlsx",
            "Gamma.xlsx",
        ]

        # All outputs are valid workbooks with template content
        for f in files:
            wb = load_workbook(str(f))
            assert wb.active["A1"].value == "Header"
            assert len(wb.sheetnames) == 1
            wb.close()

    # 11 ───────────────────────────────────────────────────────────────────
    def test_custom_sheet_and_column(self, tmp_path):
        """Config with non-default matching_sheet and filename_col."""
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))

        # Matching file with custom sheet name and column header
        wb = Workbook()
        ws = wb.active
        ws.title = "DataSheet"
        ws["A1"] = "Location"
        ws["B1"] = "PlanWork"
        ws["A2"] = "CustomSite"
        ws["B2"] = "PW999"
        wb.save(str(config["matching_file"]))
        wb.close()

        config["matching_sheet"] = "DataSheet"
        config["filename_col"] = "Location"

        main(config=config)

        out = Path(config["output_folder"])
        assert (out / "CustomSite.xlsx").exists()

    # 12 ───────────────────────────────────────────────────────────────────
    def test_missing_column_header(self, tmp_path):
        """Config filename_col missing from matching headers → SystemExit(1)."""
        config = _cfg(tmp_path)
        _make_template(Path(config["template"]))

        # Matching file has "Site" column but we ask for "NonexistentCol"
        _make_matching(Path(config["matching_file"]), [
            ("SiteA", "PW001"),
        ])

        config["filename_col"] = "NonexistentCol"

        with pytest.raises(SystemExit) as exc:
            main(config=config)
        assert exc.value.code == 1
