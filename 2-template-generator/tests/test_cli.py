"""CLI integration tests for 2-template-generator — subprocess invocation."""
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

SCRIPT = Path(__file__).parent.parent / "generate.py"


# ── helpers ──────────────────────────────────────────────────────────────

def _setup(tmp_path, config_dict=None):
    """Copy generate.py into tool2/ and optionally create config.json."""
    d = tmp_path / "tool2"
    d.mkdir()
    shutil.copy2(SCRIPT, d / "generate.py")
    if config_dict:
        (d / "config.json").write_text(json.dumps(config_dict), encoding="utf-8")
    return d


def _run(workdir, timeout=30):
    """Run generate.py with subprocess and return CompletedProcess."""
    return subprocess.run(
        [sys.executable, "generate.py"],
        cwd=str(workdir),
        capture_output=True,
        text=True,
        timeout=timeout,
    )


# ── tests ────────────────────────────────────────────────────────────────

class TestCLI:
    """Subprocess-level CLI tests for generate.py."""

    def test_happy_path(self, tmp_path):
        """Full CLI run: matching.xlsx → template copy → Done message."""
        # matching.xlsx at tmp_path level (parent of tool2/)
        wb = Workbook()
        ws = wb.active
        ws.title = "match"
        ws["A1"] = "Site"
        ws["B1"] = "PW Number"
        ws["A2"] = "TestSite"
        ws["B2"] = "XX001"
        matching_path = tmp_path / "matching.xlsx"
        wb.save(str(matching_path))
        wb.close()

        # template.xlsx inside tool2/
        d = _setup(tmp_path)
        wb2 = Workbook()
        wb2.active["A1"] = "Header"
        template_path = d / "template.xlsx"
        wb2.save(str(template_path))
        wb2.close()

        # Write config.json with relative paths matching the nested layout
        (d / "config.json").write_text(json.dumps({
            "matching_file": "../matching.xlsx",
            "matching_sheet": "match",
            "filename_col": "Site",
            "planwork_col": "PW Number",
            "template": "./template.xlsx",
            "output_folder": "./output",
        }), encoding="utf-8")

        result = _run(d)

        assert result.returncode == 0, f"stderr: {result.stderr}"
        assert "Done" in result.stdout

        out_file = d / "output" / "TestSite.xlsx"
        assert out_file.exists()
        wb_out = load_workbook(str(out_file))
        assert wb_out.active["A1"].value == "Header"
        wb_out.close()

    def test_missing_config(self, tmp_path):
        """No config.json → FileNotFoundError → non-zero exit + traceback."""
        d = _setup(tmp_path)  # no config_dict → no config.json written

        result = _run(d)

        assert result.returncode != 0
        # Python raises FileNotFoundError on open(), prints traceback to stderr
        assert "Traceback" in result.stderr
        assert "FileNotFoundError" in result.stderr

    def test_missing_template(self, tmp_path):
        """Valid config but template.xlsx missing → exit 1 + error message."""
        # matching.xlsx at tmp_path level
        wb = Workbook()
        ws = wb.active
        ws.title = "match"
        ws["A1"] = "Site"
        ws["B1"] = "PW Number"
        ws["A2"] = "TestSite"
        ws["B2"] = "XX001"
        wb.save(str(tmp_path / "matching.xlsx"))
        wb.close()

        # Setup tool2 with config but NO template.xlsx
        d = _setup(tmp_path, config_dict={
            "matching_file": "../matching.xlsx",
            "matching_sheet": "match",
            "filename_col": "Site",
            "planwork_col": "PW Number",
            "template": "./template.xlsx",
            "output_folder": "./output",
        })

        result = _run(d)

        assert result.returncode == 1
        assert "Template file not found" in result.stderr
