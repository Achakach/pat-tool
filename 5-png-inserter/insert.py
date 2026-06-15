#!/usr/bin/env python3
"""CLI for PNG inserter — match XLSX files to PNGs and insert them."""

import sys
import json
import shutil
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook
from src.matcher import read_matching, match_pngs, extract_planwork
from src.inserter import purge_sheet, extract_label, extract_site, find_matching_sheet, insert_png, insert_png_no_label, _setup_a4_print, _calc_page_rows


def progress_bar(current, total, width=30):
    filled = int(width * current / total)
    bar = "[" + "=" * filled + ">" + " " * (width - filled - 1) + "]"
    percent = int(100 * current / total)
    return f"{bar} {percent:3d}%"


def _parse_print_title_rows(value, page_rows=None):
    """Parse print_title_rows config value.

    Returns (header_count, print_title_rows_str):
        header_count: int (0 if disabled, else the end row number)
        print_title_rows_str: str or None (None if disabled)
    """
    if value is None:
        return (0, None)

    if not isinstance(value, str) or ':' not in value:
        print(f"WARNING: Invalid print_title_rows '{value}' — must be 'start:end' format. Disabled.", file=sys.stderr)
        return (0, None)

    parts = value.split(':')
    if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
        print(f"WARNING: Invalid print_title_rows '{value}' — non-numeric parts. Disabled.", file=sys.stderr)
        return (0, None)

    start, end = int(parts[0]), int(parts[1])

    if page_rows is not None:
        content_rows = page_rows - end
        if content_rows < 2:
            w = max(0, content_rows)
            print(f"WARNING: print_title_rows '{value}' leaves only {w} content row(s) per page (page_rows={page_rows}).", file=sys.stderr)

    return (end, value)


def main():
    config_path = Path(__file__).parent / "config.json"
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    matching_file = Path(config["matching_file"])
    if not matching_file.is_absolute():
        matching_file = Path(__file__).parent / matching_file

    xlsx_folder = Path(config["xlsx_folder"])
    png_folder = Path(config["png_folder"])
    output_folder = Path(config["output_folder"])

    if not xlsx_folder.is_dir():
        xlsx_folder = Path(__file__).parent / xlsx_folder
    if not png_folder.is_dir():
        png_folder = Path(__file__).parent / png_folder
    if not output_folder.is_absolute():
        output_folder = Path(__file__).parent / output_folder

    if not matching_file.exists():
        print(f"Matching file not found: {matching_file}", file=sys.stderr)
        sys.exit(1)
    if not xlsx_folder.is_dir():
        print(f"XLSX folder not found: {xlsx_folder}", file=sys.stderr)
        sys.exit(2)

    output_folder.mkdir(parents=True, exist_ok=True)

    # Read matching
    mapping = read_matching(
        str(matching_file),
        config["matching_sheet"],
        config["filename_col"],
        config["planwork_col"]
    )

    if not mapping:
        print("No entries found in matching file.", file=sys.stderr)
        sys.exit(1)

    # --- Purge config (used later with sheet matching) ---
    purge_from = config.get("purge_from_row")
    merge_to_col = config.get("label_merge_to_col")
    gap_rows = config.get("insert_gap_rows", 1)
    img_col = config.get("image_insert_col", "A")
    display_width = config.get("image_display_width")
    page_break_enabled = config.get("page_break_before_label", False)
    a4_page_rows_override = config.get("a4_page_rows")

    total_inserted = 0
    total_files = 0

    # Pre-scan: count all matchable PNGs across all files for global progress
    global_total = 0
    file_png_counts = {}
    for xlsx_path in sorted(xlsx_folder.glob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        stem = xlsx_path.stem
        if stem in mapping:
            planworks = mapping[stem]
        elif xlsx_path.name in mapping:
            planworks = mapping[xlsx_path.name]
        else:
            continue
        pngs = match_pngs(png_folder, planworks)
        if not pngs:
            continue
        valid = 0
        temp = output_folder / f"_scan_{xlsx_path.name}"
        shutil.copy2(xlsx_path, temp)
        for png in pngs:
            wb = load_workbook(str(temp))
            label = extract_label(png.name)
            sheet = find_matching_sheet(wb, label)
            wb.close()
            if sheet:
                valid += 1
        file_png_counts[xlsx_path.name] = valid
        global_total += valid
        temp.unlink()

    global_done = 0

    for xlsx_path in sorted(xlsx_folder.glob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue

        stem = xlsx_path.stem  # "one" without .xlsx
        if stem in mapping:
            planworks = mapping[stem]
        elif xlsx_path.name in mapping:
            planworks = mapping[xlsx_path.name]
        else:
            print(f"SKIP: {xlsx_path.name} — no matching entry")
            continue

        print(f"Processing: {xlsx_path.name} -> {planworks}")
        total_files += 1

        # Find matching PNGs
        pngs = match_pngs(png_folder, planworks)
        if not pngs:
            print(f"  WARNING: No PNGs matched for {planworks}")
            continue

        # Group PNGs by site so same site images stay together
        site_pngs = defaultdict(list)
        for png in pngs:
            site = extract_site(png.name)
            site_pngs[site].append(png)
        pngs = []
        for site in sorted(site_pngs):
            pngs.extend(site_pngs[site])

        # Copy XLSX to output
        output_path = output_folder / xlsx_path.name
        shutil.copy2(xlsx_path, output_path)

        # Pre-filter: only count PNGs that match a sheet
        valid_pngs = []
        for png in pngs:
            label = extract_label(png.name)
            wb = load_workbook(str(output_path))
            sheet_name = find_matching_sheet(wb, label)
            wb.close()
            if sheet_name:
                valid_pngs.append(png)

        total_pngs = len(valid_pngs)
        pngs = valid_pngs
        # Purge matched sheets + insert PNGs into sheets
        purged_sheets = set()
        sheet_page_rows = {}
        sheet_rows = {}
        labeled = set()
        inserted = 0

        for png in pngs:
            label = extract_label(png.name)

            wb = load_workbook(str(output_path))
            sheet_name = find_matching_sheet(wb, label)
            wb.close()

            if not sheet_name:
                print(f"  WARNING: No sheet matched for '{label}' ({png.name})")
                continue

            # Purge once per sheet + setup A4 print
            if purge_from and sheet_name not in purged_sheets:
                purge_sheet(output_path, sheet_name, purge_from)
                purged_sheets.add(sheet_name)
                sheet_rows[sheet_name] = purge_from
                wb = load_workbook(str(output_path))
                _setup_a4_print(wb[sheet_name])
                if page_break_enabled:
                    sheet_page_rows[sheet_name] = _calc_page_rows(wb[sheet_name], a4_page_rows_override)
                else:
                    sheet_page_rows[sheet_name] = None
                    # autoPageBreaks=True handles break management
                wb.save(str(output_path))
                wb.close()
                pr_val = sheet_page_rows.get(sheet_name)
                print(f"[DEBUG] insert.py: sheet='{sheet_name}' purged, page_rows={pr_val} (from margins T=0.5 B=0.5)", file=sys.stderr)
                print(f"  Purged: '{sheet_name}' from row {purge_from}")

            # Get current row for this sheet
            current_row = sheet_rows.get(sheet_name, purge_from or 10)

            # Insert label (site name) + PNG, or just PNG if already labeled
            site = extract_site(png.name)
            if (site, sheet_name) not in labeled:
                pr_val = sheet_page_rows.get(sheet_name)
                print(f"[DEBUG] insert.py: calling insert_png(sheet='{sheet_name}', site='{site}', current_row={current_row}, page_rows={pr_val})", file=sys.stderr)
                next_row = insert_png(output_path, sheet_name, png, site, current_row, merge_to_col, gap_rows, col=img_col, display_width=display_width, page_rows=pr_val, purge_from=(purge_from or 10))
                labeled.add((site, sheet_name))
            else:
                pr_val = sheet_page_rows.get(sheet_name)
                print(f"[DEBUG] insert.py: calling insert_png_no_label(sheet='{sheet_name}', current_row={current_row}, page_rows={pr_val})", file=sys.stderr)
                next_row = insert_png_no_label(output_path, sheet_name, png, current_row, gap_rows, col=img_col, display_width=display_width, page_rows=pr_val)

            sheet_rows[sheet_name] = next_row

            pr_val = sheet_page_rows.get(sheet_name)
            if pr_val and (next_row - current_row) > pr_val:
                print(f"  WARNING: Image '{png.name}' spans {next_row - current_row} rows, exceeding page capacity of {pr_val}", file=sys.stderr)

            pw = extract_planwork(png.name)
            gbar = progress_bar(global_done + 1, global_total)
            fbar = progress_bar(inserted + 1, file_png_counts[xlsx_path.name])
            print(f"  {gbar} | {fbar} {site} ({pw}) → '{sheet_name}'")
            inserted += 1
            global_done += 1
        total_inserted += inserted

    print(f"\nDone. Matched {total_inserted} PNGs across {total_files} files.")


if __name__ == "__main__":
    main()
