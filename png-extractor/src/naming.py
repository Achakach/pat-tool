"""Naming engine — label lookup, sanitization, and filename construction.

Provides pure string-manipulation utilities for building output filenames
from XLSX image anchor positions and cell labels.
"""

from __future__ import annotations

import re
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def col_letter(n: int) -> str:
    """Convert 0-indexed column number to Excel column letter.

    Args:
        n: 0-indexed column number (0 = A, 25 = Z, 26 = AA).

    Returns:
        Excel column letter string.
    """
    if n < 26:
        return chr(ord("A") + n)
    return col_letter(n // 26 - 1) + col_letter(n % 26)


def sanitize(name: str) -> str:
    """Sanitize a string for safe use in filenames.

    Replaces characters invalid in Windows filenames (/:*?"<>|) with
    underscores, then strips leading/trailing whitespace and dots.

    Args:
        name: Input string to sanitize.

    Returns:
        Sanitized string safe for use in filenames.
    """
    bad_chars = '/\\:*?"<>|'
    for c in bad_chars:
        name = name.replace(c, "_")
    name = name.strip()
    name = name.strip(".")
    name = name.strip()
    return name


def build_filename(
    sheet_name: str,
    label: str | None,
    anchor_row: int,
    anchor_col: int,
) -> str:
    """Build output filename for an extracted PNG.

    If *label* is truthy:
        ``{sheet}_{label}.png``
    Otherwise (None or empty):
        ``{sheet}_row{row}_col{letter}.png``

    All components are sanitized before joining.

    Args:
        sheet_name: Name of the worksheet.
        label: Label text from the cell above the image (or None/empty).
        anchor_row: 0-indexed row of the image anchor.
        anchor_col: 0-indexed column of the image anchor.

    Returns:
        Output filename string.
    """
    sheet = sanitize(sheet_name)

    if label:
        lbl = sanitize(label)
        return f"{sheet}_{lbl}.png"

    row = anchor_row + 1
    col = col_letter(anchor_col)
    return f"{sheet}_row{row}_col{col}.png"


def get_label(ws, anchor_row: int, anchor_col: int) -> str | None:
    """Scan upward from the cell above an anchor to find a label.

    Search order for each row going upward:
    1. Same column above anchor
    2. Any other column in that row (left to right)
    3. When all rows above exhausted: check the anchor cell itself
    4. If nothing found anywhere: return None (fallback naming)

    Args:
        ws: openpyxl worksheet object.
        anchor_row: 0-indexed row of the image anchor.
        anchor_col: 0-indexed column of the image anchor.

    Returns:
        Stripped cell value if a non-empty cell is found, else None.

    Notes:
        - If *anchor_row* is 0 the upward loop is skipped but the anchor
          cell itself is still checked.
        - openpyxl uses 1-indexed cells; the column lookup adds 1 to
          *anchor_col*.
    """
    max_col = ws.max_column or 26

    if anchor_row > 0:
        for row in range(anchor_row, 0, -1):
            # Step 1: same column
            value = ws.cell(row=row, column=anchor_col + 1).value
            if value and str(value).strip():
                return str(value).strip()

            # Step 2: scan entire row (left to right), skip already-checked column
            for col in range(1, max_col + 1):
                if col == anchor_col + 1:
                    continue
                value = ws.cell(row=row, column=col).value
                if value and str(value).strip():
                    return str(value).strip()

    # Step 3: check the anchor cell itself
    anchor_value = ws.cell(row=anchor_row + 1, column=anchor_col + 1).value
    if anchor_value and str(anchor_value).strip():
        return str(anchor_value).strip()

    return None


def parse_prefix(sheet_name: str) -> tuple[str, str] | None:
    """Extract (prefix, site) from 'exist bkk101' or 'new bkk999'. None if no match."""
    m = re.match(r'(exist|new)\s+(.+)', sheet_name, re.IGNORECASE)
    if m:
        return (m.group(1).lower(), m.group(2).strip())
    return None


def get_label_with_row(ws, anchor_row: int, anchor_col: int) -> tuple[str, int] | None:
    """Like get_label() but returns (text, row_found) or None.

    row_found is 1-indexed openpyxl row where text was found.
    """
    max_col = ws.max_column or 26
    if anchor_row > 0:
        for row in range(anchor_row, 0, -1):
            value = ws.cell(row=row, column=anchor_col + 1).value
            if value and str(value).strip():
                return (str(value).strip(), row)
            for col in range(1, max_col + 1):
                if col == anchor_col + 1:
                    continue
                value = ws.cell(row=row, column=col).value
                if value and str(value).strip():
                    return (str(value).strip(), row)
    anchor_value = ws.cell(row=anchor_row + 1, column=anchor_col + 1).value
    if anchor_value and str(anchor_value).strip():
        return (str(anchor_value).strip(), anchor_row + 1)
    return None


def build_pw_filename(planwork: str, prefix: str, site: str, label_text: str) -> str:
    """Build output filename: PW {planwork}_{prefix} {site}_{label_text}.png"""
    return f"PW {planwork}_{prefix} {site}_{sanitize(label_text)}.png"
