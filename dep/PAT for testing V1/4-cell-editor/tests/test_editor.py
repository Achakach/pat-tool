"""Tests for cell-editor."""
import sys
from pathlib import Path
import pytest
from openpyxl import Workbook
from openpyxl import load_workbook
from src.editor import process_workbook
from edit import main


class TestProcessWorkbook:

    def test_replaces_right_cell(self, tmp_path):
        """Label cell untouched, right cell gets the replacement."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name:"
        ws["B1"] = "old_value"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"name:": "kacha"}, "first")
        assert changed == 1
        
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["A1"].value == "name:"        # untouched
        assert ws2["B1"].value == "kacha"         # replaced
        wb2.close()

    def test_merged_cell_skips_right(self, tmp_path):
        """A3:B3 merged → right of A3 is C3 (skips B3)."""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A3:B3")
        ws["A3"] = "surname:"
        ws["C3"] = "old"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"surname:": "picolo"}, "first")
        assert changed == 1
        
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["A3"].value == "surname:"   # untouched
        assert ws2["C3"].value == "picolo"      # replaced (skipped B3)
        wb2.close()

    def test_match_mode_first(self, tmp_path):
        """Only first occurrence replaced, second ignored."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name:"
        ws["B1"] = "first"
        ws["A5"] = "name:"
        ws["B5"] = "second"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"name:": "kacha"}, "first")
        assert changed == 1
        
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["B1"].value == "kacha"       # first → replaced
        assert ws2["B5"].value == "second"       # second → untouched
        wb2.close()

    def test_match_mode_all(self, tmp_path):
        """All occurrences replaced."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name:"
        ws["B1"] = "first"
        ws["A5"] = "name:"
        ws["B5"] = "second"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"name:": "kacha"}, "all")
        assert changed == 2
        
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["B1"].value == "kacha"
        assert ws2["B5"].value == "kacha"
        wb2.close()

    def test_no_match_unchanged(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "untouched"
        ws["B1"] = "keep"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"name:": "kacha"}, "first")
        assert changed == 0
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp_path / "out.xlsx")
        assert wb2.active["B1"].value == "keep"
        wb2.close()

    # ── New tests (Task 2) ──────────────────────────────────────────

    def test_multiple_prefixes(self, tmp_path):
        """3 prefixes in replacements → all 3 right-side cells replaced."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name:"
        ws["B1"] = "old_name"
        ws["A2"] = "date:"
        ws["B2"] = "old_date"
        ws["A3"] = "amount:"
        ws["B3"] = "old_amount"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        replacements = {"name:": "kacha", "date:": "2024-01-01", "amount:": "99.99"}
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   replacements, "first")
        assert changed == 3

        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["A1"].value == "name:"       # untouched
        assert ws2["B1"].value == "kacha"
        assert ws2["A2"].value == "date:"       # untouched
        assert ws2["B2"].value == "2024-01-01"
        assert ws2["A3"].value == "amount:"     # untouched
        assert ws2["B3"].value == "99.99"
        wb2.close()

    def test_multi_worksheet(self, tmp_path):
        """Replacements applied in both worksheets."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws1["A1"] = "name:"
        ws1["B1"] = "old1"
        ws2["A1"] = "name:"
        ws2["B1"] = "old2"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"name:": "kacha"}, "all")
        assert changed == 2

        wb2 = load_workbook(tmp_path / "out.xlsx")
        assert wb2["Sheet1"]["B1"].value == "kacha"
        assert wb2["Sheet2"]["B1"].value == "kacha"
        wb2.close()

    def test_thai_unicode_replacements(self, tmp_path):
        """Thai prefix + Thai replacement — Unicode round-trip works."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "ชื่อ"
        ws["B1"] = "ค่าเก่า"
        ws["A2"] = "วันที่"
        ws["B2"] = "31/12/2567"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        replacements = {"ชื่อ": "กชกร", "วันที่": "1/1/2568"}
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   replacements, "first")
        assert changed == 2

        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["A1"].value == "ชื่อ"         # untouched
        assert ws2["B1"].value == "กชกร"
        assert ws2["A2"].value == "วันที่"       # untouched
        assert ws2["B2"].value == "1/1/2568"
        wb2.close()

    def test_overlapping_prefixes(self, tmp_path):
        """Longer prefix checked first → no false match on shorter prefix."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name:"
        ws["B1"] = "short_old"
        ws["A2"] = "name_detailed:"
        ws["B2"] = "detailed_old"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        # Longer prefix first so "name_detailed:" not stolen by "name:"
        replacements = {"name_detailed:": "detailed_new", "name:": "short_new"}
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   replacements, "first")
        assert changed == 2

        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["B1"].value == "short_new"       # "name:" → short_new
        assert ws2["B2"].value == "detailed_new"    # "name_detailed:" → detailed_new
        wb2.close()


class TestMainIntegration:
    """Integration tests for edit.main() with config dict injection."""

    def test_main_with_config_dict(self, tmp_path):
        """Full integration: create input xlsx, run main(config=…), verify output."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        output_dir = tmp_path / "output"

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "old:"
        ws["B1"] = "replace_me"
        wb.save(str(input_dir / "test.xlsx"))
        wb.close()

        config = {
            "input_folder": str(input_dir),
            "output_folder": str(output_dir),
            "replacements": {"old:": "new_val"},
            "match_mode": "first",
        }
        main(config=config)

        wb2 = load_workbook(output_dir / "test.xlsx")
        assert wb2.active["A1"].value == "old:"        # untouched
        assert wb2.active["B1"].value == "new_val"     # replaced
        wb2.close()

    def test_main_empty_replacements(self, tmp_path):
        """Empty replacements → SystemExit(1)."""
        config = {
            "input_folder": str(tmp_path / "input"),
            "output_folder": str(tmp_path / "output"),
            "replacements": {},
        }
        with pytest.raises(SystemExit) as exc:
            main(config=config)
        assert exc.value.code == 1

    def test_main_missing_input_folder(self, tmp_path):
        """Non-existent input folder → SystemExit(2)."""
        nonexistent = str(tmp_path / "nonexistent_folder_xyz")
        config = {
            "input_folder": nonexistent,
            "output_folder": str(tmp_path / "output"),
            "replacements": {"x:": "y"},
        }
        with pytest.raises(SystemExit) as exc:
            main(config=config)
        assert exc.value.code == 2
