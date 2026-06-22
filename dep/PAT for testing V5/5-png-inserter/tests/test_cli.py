"""CLI integration tests for 5-png-inserter — subprocess invocation."""
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook

SCRIPT = Path(__file__).parent.parent / "insert.py"
SRC = Path(__file__).parent.parent / "src"

# Minimal valid 1×1 pixel PNG
_PNG_DATA = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
    b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00"
    b"\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82"
)

_CFG_BASE = {
    "matching_file": "../matching.xlsx",
    "matching_sheet": "match",
    "filename_col": "Site",
    "planwork_col": "PW Number",
    "xlsx_folder": "./xlsx",
    "png_folder": "./input",
    "output_folder": "./out",
    "purge_from_row": 10,
    "label_merge_to_col": "K",
    "insert_gap_rows": 1,
    "image_insert_col": "C",
    "page_break_before_label": False,
    "print_title_rows": None,
    "a4_page_rows": None,
    "image_display_width": 100,
}


def _setup(tmp_path, config_dict=None):
    d = tmp_path / "tool5"
    d.mkdir()
    shutil.copy2(SCRIPT, d / "insert.py")
    shutil.copytree(SRC, d / "src")
    if config_dict:
        (d / "config.json").write_text(json.dumps(config_dict), encoding="utf-8")
    return d


class TestCLI:
    """Subprocess integration tests for the insert.py CLI."""

    def test_happy_path(self, tmp_path):
        """Full CLI run: matching + PNG + XLSX → image inserted → Done."""
        # matching.xlsx at tmp_path level (parent of tool5/)
        wb = Workbook()
        ws = wb.active
        ws.title = "match"
        ws["A1"] = "Site"
        ws["B1"] = "PW Number"
        ws["A2"] = "TestSite"
        ws["B2"] = "XX001"
        wb.save(str(tmp_path / "matching.xlsx"))
        wb.close()

        d = _setup(tmp_path)

        # input/ directory with PNG
        in_dir = d / "input"
        in_dir.mkdir()
        (in_dir / "PW XX001_exist TestSite_label.png").write_bytes(_PNG_DATA)

        # xlsx/ directory with XLSX (12+ rows so purge doesn't break)
        xlsx_dir = d / "xlsx"
        xlsx_dir.mkdir()
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "label"
        for i in range(1, 13):
            ws2.cell(row=i, column=1).value = f"row{i}"
        wb2.save(str(xlsx_dir / "TestSite.xlsx"))
        wb2.close()

        # config
        (d / "config.json").write_text(json.dumps(_CFG_BASE), encoding="utf-8")

        result = subprocess.run(
            [sys.executable, "insert.py"],
            cwd=str(d),
            capture_output=True,
            text=True,
            timeout=30,
        )

        assert result.returncode == 0, (
            f"exit {result.returncode}\nstderr: {result.stderr}\nstdout: {result.stdout}"
        )
        assert "Done" in result.stdout

        out_file = d / "out" / "TestSite.xlsx"
        assert out_file.exists()

    def test_missing_config(self, tmp_path):
        """No config.json → FileNotFoundError → non-zero exit."""
        d = _setup(tmp_path)  # no config_dict → no config.json

        result = subprocess.run(
            [sys.executable, "insert.py"],
            cwd=str(d),
            capture_output=True,
            text=True,
            timeout=30,
        )

        assert result.returncode != 0

    def test_missing_xlsx_folder(self, tmp_path):
        """Valid config but missing xlsx/ → exit 2 + error message."""
        # matching.xlsx required to pass first guard (exit 1 otherwise)
        wb = Workbook()
        ws = wb.active
        ws.title = "match"
        ws["A1"] = "Site"
        ws["B1"] = "PW Number"
        ws["A2"] = "TestSite"
        ws["B2"] = "XX001"
        wb.save(str(tmp_path / "matching.xlsx"))
        wb.close()

        d = _setup(tmp_path, config_dict=_CFG_BASE)
        # Do NOT create xlsx/ directory

        result = subprocess.run(
            [sys.executable, "insert.py"],
            cwd=str(d),
            capture_output=True,
            text=True,
            timeout=30,
        )

        assert result.returncode == 2, (
            f"Expected exit 2, got {result.returncode}\nstderr: {result.stderr}"
        )
        assert "XLSX folder not found" in result.stderr

    def test_happy_path_persistent(self):
        """Full CLI run — output left in tests/cli_output/ for manual inspection."""
        base = Path(__file__).parent / "cli_work"
        if base.exists():
            shutil.rmtree(base)

        d = base / "tool5"
        d.mkdir(parents=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "match"
        ws["A1"] = "Site"
        ws["B1"] = "PW Number"
        ws["A2"] = "TestSite"
        ws["B2"] = "XX001"
        wb.save(str(base / "matching.xlsx"))
        wb.close()

        shutil.copy2(SCRIPT, d / "insert.py")
        shutil.copytree(SRC, d / "src")

        in_dir = d / "input"
        in_dir.mkdir()
        (in_dir / "PW XX001_exist TestSite_label.png").write_bytes(_PNG_DATA)

        xlsx_dir = d / "xlsx"
        xlsx_dir.mkdir()
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "label"
        for i in range(1, 13):
            ws2.cell(row=i, column=1).value = f"row{i}"
        wb2.save(str(xlsx_dir / "TestSite.xlsx"))
        wb2.close()

        (d / "config.json").write_text(json.dumps(_CFG_BASE), encoding="utf-8")

        result = subprocess.run(
            [sys.executable, "insert.py"],
            cwd=str(d),
            capture_output=True,
            text=True,
            timeout=30,
        )

        assert result.returncode == 0, (
            f"exit {result.returncode}\nstderr: {result.stderr}"
        )
        assert "Done" in result.stdout

        out_dir = Path(__file__).parent / "cli_output"
        if out_dir.exists():
            shutil.rmtree(out_dir)
        out_dir.mkdir()
        for f in (d / "out").glob("*"):
            shutil.copy2(f, out_dir / f.name)

        print(f"\n  Persistent output: {out_dir.resolve()}")
        for f in sorted(out_dir.glob("*")):
            print(f"    {f.name}")
