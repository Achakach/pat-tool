"""Tests for cell-editor."""
import sys
from pathlib import Path
import pytest
from openpyxl import Workbook
from src.editor import process_workbook


class TestProcessWorkbook:

    def test_replaces_right_cell(self, tmp_path):
        """Label cell untouched, right cell gets the replacement."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name:"
        ws["B1"] = "old_value"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"name:": "kacha"}, "first")
        assert changed == 1
        
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["A1"].value == "name:"        # untouched
        assert ws2["B1"].value == "kacha"         # replaced
        wb2.close()

    def test_merged_cell_skips_right(self, tmp_path):
        """A3:B3 merged → right of A3 is C3 (skips B3)."""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A3:B3")
        ws["A3"] = "surname:"
        ws["C3"] = "old"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"surname:": "picolo"}, "first")
        assert changed == 1
        
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["A3"].value == "surname:"   # untouched
        assert ws2["C3"].value == "picolo"      # replaced (skipped B3)
        wb2.close()

    def test_match_mode_first(self, tmp_path):
        """Only first occurrence replaced, second ignored."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name:"
        ws["B1"] = "first"
        ws["A5"] = "name:"
        ws["B5"] = "second"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"name:": "kacha"}, "first")
        assert changed == 1
        
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["B1"].value == "kacha"       # first → replaced
        assert ws2["B5"].value == "second"       # second → untouched
        wb2.close()

    def test_match_mode_all(self, tmp_path):
        """All occurrences replaced."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name:"
        ws["B1"] = "first"
        ws["A5"] = "name:"
        ws["B5"] = "second"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"name:": "kacha"}, "all")
        assert changed == 2
        
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp_path / "out.xlsx")
        ws2 = wb2.active
        assert ws2["B1"].value == "kacha"
        assert ws2["B5"].value == "kacha"
        wb2.close()

    def test_no_match_unchanged(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "untouched"
        ws["B1"] = "keep"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()
        
        changed = process_workbook(xlsx_path, tmp_path / "out.xlsx",
                                   {"name:": "kacha"}, "first")
        assert changed == 0
        from openpyxl import load_workbook
        wb2 = load_workbook(tmp_path / "out.xlsx")
        assert wb2.active["B1"].value == "keep"
        wb2.close()
