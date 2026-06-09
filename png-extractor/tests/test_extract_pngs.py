"""Pytest test suite for PAT tool — naming, config, and CLI integration.

Covers:
- naming.py: col_letter, sanitize, build_filename, get_label
- config.py: load_config validation
- extract_pngs.py: CLI orchestration via subprocess
"""

from __future__ import annotations

import pytest
import subprocess
import shutil
import sys
import json
from pathlib import Path
from unittest.mock import MagicMock

from src.naming import (
    col_letter, sanitize, build_filename, get_label,
    parse_prefix, get_label_with_row, build_pw_filename,
)
from src.config import load_config

# ── Paths ─────────────────────────────────────────────────────────────────

FIXTURE = Path(__file__).parent.parent.parent / "test_fixture.xlsx"
EXTRACT_SCRIPT = Path(__file__).parent.parent / "extract_pngs.py"


# ── TestNaming ────────────────────────────────────────────────────────────


class TestNaming:
    """Unit tests for naming.py functions."""

    @pytest.mark.parametrize(
        "n, expected",
        [
            (0, "A"),
            (1, "B"),
            (25, "Z"),
            (26, "AA"),
            (27, "AB"),
            (51, "AZ"),
            (52, "BA"),
        ],
    )
    def test_col_letter(self, n, expected):
        assert col_letter(n) == expected

    def test_sanitize_bad_chars(self):
        result = sanitize('test/name:with*bad?chars')
        assert result == 'test_name_with_bad_chars'

    def test_sanitize_bad_chars_all(self):
        """Verify all /:*?"<>| become underscores."""
        result = sanitize('a/b:c*d?e"f<g>h|i')
        assert result == 'a_b_c_d_e_f_g_h_i'

    def test_sanitize_edges(self):
        result = sanitize('.  hello  .')
        assert result == 'hello'

    def test_sanitize_noop(self):
        result = sanitize('clean_name')
        assert result == 'clean_name'

    def test_build_filename_with_label(self):
        result = build_filename("Sheet1", "Revenue", 3, 1)
        assert result == "Sheet1_Revenue.png"

    def test_build_filename_fallback(self):
        result = build_filename("Sheet1", None, 3, 1)
        assert result == "Sheet1_row4_colB.png"

    def test_build_filename_row_zero(self):
        result = build_filename("Sheet1", "", 0, 0)
        assert result == "Sheet1_row1_colA.png"

    def test_build_filename_sanitizes_all(self):
        result = build_filename("Q1/2024", "Profit/Loss", 1, 1)
        assert result == "Q1_2024_Profit_Loss.png"

    def test_build_filename_large_col(self):
        """Column >= 26 should produce multi-letter (e.g. AA)."""
        result = build_filename("S1", None, 0, 26)
        assert result == "S1_row1_colAA.png"

    def test_get_label_row_zero(self):
        ws = MagicMock()
        ws.cell.return_value.value = None  # anchor cell empty
        result = get_label(ws, 0, 0)
        assert result is None

    def test_get_label_with_value(self):
        ws = MagicMock()
        ws.cell.return_value.value = "Hello"
        result = get_label(ws, 2, 1)
        assert result == "Hello"

    def test_get_label_whitespace_only(self):
        """Cell with only whitespace should return None."""
        ws = MagicMock()
        ws.cell.return_value.value = "   "
        ws.max_column = 5
        result = get_label(ws, 1, 0)
        assert result is None

    def test_get_label_numeric_value(self):
        """Numeric cell value should be cast to string."""
        ws = MagicMock()
        ws.cell.return_value.value = 42
        result = get_label(ws, 1, 0)
        assert result == "42"

    def test_get_label_upward_search(self):
        """When the immediate row above is empty, scan upward."""
        ws = MagicMock()

        def cell_side_effect(row, column):
            mock = MagicMock()
            if row == 3:
                mock.value = None
            elif row == 2:
                mock.value = "Found"
            elif row == 1:
                mock.value = None
            return mock

        ws.cell.side_effect = cell_side_effect
        ws.max_column = 5
        result = get_label(ws, 3, 0)  # anchor at row 3 (0-indexed)
        assert result == "Found"

    def test_get_label_adjacent_column(self):
        """Label in a different column but same row above the image."""
        ws = MagicMock()
        row_values = {2: None, 3: "Nearby Label"}  # column 2 empty, column 3 has text

        def cell_side_effect(row, column):
            mock = MagicMock()
            if row == 2:  # row above anchor
                mock.value = row_values.get(column)
            return mock

        ws.cell.side_effect = cell_side_effect
        ws.max_column = 5
        result = get_label(ws, 2, 1)  # anchor at row=2, col=1 (0-indexed)
        assert result == "Nearby Label"

    def test_get_label_anchor_cell(self):
        """Label is in the same cell as the image anchor."""
        ws = MagicMock()

        def cell_side_effect(row, column):
            mock = MagicMock()
            if row == 2 and column == 2:  # anchor cell (openpyxl row 2, col 2)
                mock.value = "Inline Label"
            else:
                mock.value = None
            return mock

        ws.cell.side_effect = cell_side_effect
        ws.max_column = 5
        result = get_label(ws, 1, 1)  # anchor at row=1, col=1 (0-indexed)
        assert result == "Inline Label"


    # ── New naming tests ─────────────────────────────────────────────────

    def test_parse_prefix_exist(self):
        result = parse_prefix("exist bkk101")
        assert result == ("exist", "bkk101")

    def test_parse_prefix_new(self):
        result = parse_prefix("New BKK999")
        assert result == ("new", "BKK999")

    def test_parse_prefix_none(self):
        assert parse_prefix("Summary") is None
        assert parse_prefix("PW bkk007") is None

    def test_parse_prefix_edge_cases(self):
        # "exist"/"new" alone — no whitespace, no site — returns None
        assert parse_prefix("exist") is None
        assert parse_prefix("new") is None
        assert parse_prefix("EXIST site_a") == ("exist", "site_a")

    def test_build_pw_filename(self):
        result = build_pw_filename("bkk007", "exist", "bkk101", "My Label")
        assert result == "PW bkk007_exist bkk101_My Label.png"

    def test_build_pw_filename_new(self):
        result = build_pw_filename("bkk007", "new", "bkk999", "Test Label")
        assert result == "PW bkk007_new bkk999_Test Label.png"

    def test_get_label_with_row(self):
        ws = MagicMock()
        ws.cell.return_value.value = "Hello"
        result = get_label_with_row(ws, 2, 1)
        # anchor_row=2 (0-indexed) → openpyxl anchor row is 3
        # label found at openpyxl row 2 (one row above anchor)
        assert result == ("Hello", 2)

    def test_get_label_with_row_none(self):
        ws = MagicMock()
        ws.cell.return_value.value = None
        result = get_label_with_row(ws, 0, 0)
        assert result is None


# ── TestConfig ────────────────────────────────────────────────────────────


class TestConfig:
    """Unit tests for config.py load_config."""

    def test_load_valid_config(self, tmp_path):
        config_file = tmp_path / "config.json"
        config_file.write_text('{"input_folder": "./in", "output_folder": "./out"}')
        result = load_config(str(config_file))
        assert result == {"input_folder": "./in", "output_folder": "./out"}

    def test_load_missing_config(self):
        with pytest.raises(SystemExit) as exc_info:
            load_config("nonexistent.json")
        assert exc_info.value.code == 1

    def test_load_invalid_schema(self, tmp_path):
        config_file = tmp_path / "config.json"
        config_file.write_text('{"output_folder": "./out"}')
        with pytest.raises(SystemExit) as exc_info:
            load_config(str(config_file))
        assert exc_info.value.code == 1

    def test_load_non_dict(self, tmp_path):
        """JSON that parses to a list instead of dict should exit 1."""
        config_file = tmp_path / "config.json"
        config_file.write_text('[1, 2, 3]')
        with pytest.raises(SystemExit) as exc_info:
            load_config(str(config_file))
        assert exc_info.value.code == 1

    def test_load_invalid_json(self, tmp_path):
        """Malformed JSON should exit 1."""
        config_file = tmp_path / "config.json"
        config_file.write_text('{not valid json')
        with pytest.raises(SystemExit) as exc_info:
            load_config(str(config_file))
        assert exc_info.value.code == 1


# ── TestIntegration ───────────────────────────────────────────────────────


class TestIntegration:
    """Integration tests for the full extract_pngs.py pipeline."""

    def test_extract_fixture_output_new_naming(self, tmp_path):
        """Integration test: new naming format with PW/exist/new sheets.

        Uses the real fixture but renames sheets to match the new convention.
        Sheet names: "PW test", "exist site1", "new site2", "Summary".
        """
        import openpyxl

        # ── Setup temp directories ─────────────────────────────────────
        in_dir = tmp_path / "input"
        out_dir = tmp_path / "output"
        in_dir.mkdir()

        # Copy fixture into input dir
        fixture_copy = in_dir / "test_fixture.xlsx"
        shutil.copy2(FIXTURE, fixture_copy)

        # Rename sheets to match new convention
        wb = openpyxl.load_workbook(fixture_copy)
        wb.worksheets[0].title = "PW test"          # planwork sheet (skipped by prefix filter)
        wb.worksheets[1].title = "exist site1"       # extracted with "exist" prefix
        wb.worksheets[2].title = "new site2"         # extracted with "new" prefix
        wb.create_sheet("Summary")                   # skipped (no exist/new prefix)
        wb.save(str(fixture_copy))
        wb.close()

        # Create config with absolute paths
        config_file = tmp_path / "config.json"
        config_file.write_text(
            '{{"input_folder": "{}", "output_folder": "{}", "noise_threshold": 500}}'.format(
                in_dir.as_posix(), out_dir.as_posix()
            )
        )

        # Backup real config, copy test config to script dir, restore after
        script_config = EXTRACT_SCRIPT.parent / "config.json"
        backup = script_config.with_suffix(".json.bak")
        if backup.exists():
            backup.unlink()
        if script_config.exists():
            script_config.rename(backup)
        try:
            shutil.copy2(str(config_file), str(script_config))

            # ── Run CLI ────────────────────────────────────────────────────
            result = subprocess.run(
                [sys.executable, str(EXTRACT_SCRIPT)],
                capture_output=True,
                text=True,
            )

            assert result.returncode == 0, (
                f"CLI exit code {result.returncode}\n"
                f"STDERR: {result.stderr}\n"
                f"STDOUT: {result.stdout}"
            )

            # ── Verify output files match new naming format ────────────────
            # Fixture has labels only in "Edge" (now "new site2") sheet:
            #   B10="Deep Label" → extracted as PW test_new site2_row10.png
            # Images without labels are skipped (new behavior).
            output_files = sorted(out_dir.iterdir())
            assert len(output_files) == 1, (
                f"Expected 1 output file, got {len(output_files)}: "
                f"{[f.name for f in output_files]}"
            )

            fpath = output_files[0]
            name = fpath.name
            assert fpath.stat().st_size > 0, f"Empty output file: {name}"
            assert name == "PW test_new site2_Deep Label.png", (
                f"Unexpected filename: {name}"
            )
        finally:
            if script_config.exists():
                script_config.unlink()
            if backup.exists():
                backup.rename(script_config)

    def test_missing_config_exit_code(self):
        config_path = EXTRACT_SCRIPT.parent / "config.json"
        backup = config_path.with_suffix(".json.bak")
        if backup.exists():
            backup.unlink()
        if config_path.exists():
            config_path.rename(backup)
        try:
            result = subprocess.run(
                [sys.executable, str(EXTRACT_SCRIPT)],
                capture_output=True,
                text=True,
            )
            assert result.returncode == 1
        finally:
            if backup.exists():
                backup.rename(config_path)

    def test_bad_input_folder_exit_code(self, tmp_path):
        real_config = EXTRACT_SCRIPT.parent / "config.json"
        backup = real_config.with_suffix(".json.bak")
        if backup.exists():
            backup.unlink()
        if real_config.exists():
            real_config.rename(backup)
        try:
            bad_config = {"input_folder": str(tmp_path / "nonexistent"), "output_folder": str(tmp_path / "output")}
            real_config.write_text(json.dumps(bad_config))
            result = subprocess.run(
                [sys.executable, str(EXTRACT_SCRIPT)],
                capture_output=True,
                text=True,
            )
            assert result.returncode == 2
        finally:
            if real_config.exists():
                real_config.unlink()
            if backup.exists():
                backup.rename(real_config)
