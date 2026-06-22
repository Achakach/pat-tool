"""Naming engine — label lookup, sanitization, and filename construction.

Provides pure string-manipulation utilities for building output filenames
from XLSX image anchor positions and cell labels.
"""

from __future__ import annotations

import re
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def col_letter(n: int) -> str:
    """Convert 0-indexed column number to Excel column letter.

    Args:
        n: 0-indexed column number (0 = A, 25 = Z, 26 = AA).

    Returns:
        Excel column letter string.
    """
    if n < 26:
        return chr(ord("A") + n)
    return col_letter(n // 26 - 1) + col_letter(n % 26)


def sanitize(name: str) -> str:
    """Sanitize a string for safe use in filenames.

    Replaces characters invalid in Windows filenames (/:*?"<>|) with
    underscores, then strips leading/trailing whitespace and dots.

    Args:
        name: Input string to sanitize.

    Returns:
        Sanitized string safe for use in filenames.
    """
    bad_chars = '/\\:*?"<>|'
    for c in bad_chars:
        name = name.replace(c, "_")
    name = name.strip()
    name = name.strip(".")
    name = name.strip()
    return name


def parse_prefix(sheet_name: str) -> tuple[str, str] | None:
    """Extract (prefix, site) from 'exist bkk101' or 'new bkk999'. None if no match."""
    m = re.match(r'(exist|new)\s+(.+)', sheet_name, re.IGNORECASE)
    if m:
        return (m.group(1).lower(), m.group(2).strip())
    return None


def get_label_with_row(ws, anchor_row: int, anchor_col: int) -> tuple[str, int] | None:
    """Like get_label() but returns (text, row_found) or None.

    row_found is 1-indexed openpyxl row where text was found.
    """
    max_col = ws.max_column or 26
    if anchor_row > 0:
        for row in range(anchor_row, 0, -1):
            value = ws.cell(row=row, column=anchor_col + 1).value
            if value and str(value).strip():
                return (str(value).strip(), row)
            for col in range(1, max_col + 1):
                if col == anchor_col + 1:
                    continue
                value = ws.cell(row=row, column=col).value
                if value and str(value).strip():
                    return (str(value).strip(), row)
    anchor_value = ws.cell(row=anchor_row + 1, column=anchor_col + 1).value
    if anchor_value and str(anchor_value).strip():
        return (str(anchor_value).strip(), anchor_row + 1)
    return None


def build_pw_filename(planwork: str, prefix: str, site: str, label_text: str) -> str:
    """Build output filename: PW {planwork}_{prefix} {site}_{label_text}.png"""
    return f"PW {planwork}_{prefix} {site}_{sanitize(label_text)}.png"
