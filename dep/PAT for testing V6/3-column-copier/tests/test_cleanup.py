"""Tests for cleanup action — deleting build_at/paste_to columns from source XLSX."""
import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
from copier import main


def _make_matching(path, sheet="Sheet1"):
    """Create minimal matching.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws["A1"] = "Site"
    ws["B1"] = "PW Number"
    ws["A2"] = "dummy"
    ws["B2"] = "XX001"
    wb.save(str(path))
    wb.close()


def _make_source(path, data_sheet="Cutsheet", cols=19, rows=4):
    """Create source XLSX with columns A through col_letter populated.

    cols: number of columns (e.g. 19 = A-S, rendered as integer count)
    rows: number of data rows (including header row 1)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = data_sheet
    for c in range(1, cols + 1):
        for r in range(1, rows + 1):
            ws.cell(row=r, column=c).value = f"R{r}C{c}"
    wb.save(str(path))
    wb.close()


class TestCleanup:
    """Tests for action="cleanup" deleting build_at columns."""

    def test_cleanup_deletes_build_at_columns(self, tmp_path):
        """Cleanup deletes Q, R, S columns (build_at). Column count drops by 3."""
        # --- setup ---
        src_dir = tmp_path / "source"
        src_dir.mkdir()
        out_dir = tmp_path / "output"
        matching_path = tmp_path / "matching.xlsx"
        _make_matching(matching_path)

        src_path = src_dir / "test.xlsx"
        _make_source(src_path, cols=21, rows=4)  # A-U populated

        # Count columns before
        wb_before = load_workbook(str(src_path))
        cols_before = wb_before["Cutsheet"].max_column
        wb_before.close()

        config = {
            "action": "cleanup",
            "matching_file": str(matching_path),
            "source_folder": str(src_dir),
            "target_folder": str(tmp_path / "target"),
            "output_folder": str(out_dir),
            "data_sheet": "Cutsheet",
            "target_sheet": "IP",
            "matching_sheet": "Sheet1",
            "filename_col": "Site",
            "planwork_col": "PW Number",
            "source_start_row": 2,
            "paste_start_row": 3,
            "columns": {
                "pw": {"type": "planwork", "build_at": "Q", "paste_to": "J"},
                "exist_ip": {"type": "ip_lookup", "build_at": "R", "paste_to": "E", "lookup_col": "A", "log_sheet": "Log"},
                "new_ip": {"type": "ip_lookup", "build_at": "S", "paste_to": "H", "lookup_col": "B", "log_sheet": "Log"},
            },
            "print_title_rows": "1:6",
            "page_break_enabled": False,
        }

        # --- run ---
        main(config=config)

        # --- verify ---
        out_path = out_dir / "test.xlsx"
        assert out_path.exists(), "Output file not created"

        wb_after = load_workbook(str(out_path))
        ws_after = wb_after["Cutsheet"]
        cols_after = ws_after.max_column

        # 3 columns (build_at: Q/R/S) deleted
        assert cols_after == cols_before - 3, \
            f"Expected {cols_before - 3} columns, got {cols_after}"

        # Column A data intact (early columns unaffected by right-side deletes)
        assert ws_after.cell(row=1, column=1).value == "R1C1"
        assert ws_after.cell(row=2, column=1).value == "R2C1"

        wb_after.close()

    def test_cleanup_preserves_data_columns(self, tmp_path):
        """After cleanup, data columns C, D, E, G, H retain original values."""
        # --- setup ---
        src_dir = tmp_path / "source"
        src_dir.mkdir()
        out_dir = tmp_path / "output"
        matching_path = tmp_path / "matching.xlsx"
        _make_matching(matching_path)

        src_path = src_dir / "data_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Cutsheet"
        # Row 1 = header
        ws["A1"] = "HdrA"; ws["B1"] = "HdrB"; ws["C1"] = "NE_NO1"
        ws["D1"] = "Port1"; ws["E1"] = "Label1"; ws["F1"] = "HdrF"
        ws["G1"] = "NE_NO2"; ws["H1"] = "Port2"; ws["I1"] = "HdrI"
        # Fill early columns up through U so sequential deletes don't error
        for c in range(1, 22):  # A-U
            if ws.cell(row=1, column=c).value is None:
                ws.cell(row=1, column=c).value = f"Hdr{c}"
        # Data rows 2-4
        for r in range(2, 5):
            ws.cell(row=r, column=3).value = f"NE1_R{r}"   # C
            ws.cell(row=r, column=4).value = f"PORT1_R{r}"  # D
            ws.cell(row=r, column=5).value = f"LBL1_R{r}"   # E
            ws.cell(row=r, column=7).value = f"NE2_R{r}"    # G
            ws.cell(row=r, column=8).value = f"PORT2_R{r}"  # H
            for c in range(1, 22):
                if ws.cell(row=r, column=c).value is None:
                    ws.cell(row=r, column=c).value = f"R{r}C{c}"
        wb.save(str(src_path))
        wb.close()

        config = {
            "action": "cleanup",
            "matching_file": str(matching_path),
            "source_folder": str(src_dir),
            "target_folder": str(tmp_path / "target"),
            "output_folder": str(out_dir),
            "data_sheet": "Cutsheet",
            "target_sheet": "IP",
            "matching_sheet": "Sheet1",
            "filename_col": "Site",
            "planwork_col": "PW Number",
            "source_start_row": 2,
            "paste_start_row": 3,
            "columns": {
                "pw": {"type": "planwork", "build_at": "Q", "paste_to": "J"},
                "exist_ip": {"type": "ip_lookup", "build_at": "R", "paste_to": "E", "lookup_col": "A", "log_sheet": "Log"},
                "new_ip": {"type": "ip_lookup", "build_at": "S", "paste_to": "H", "lookup_col": "B", "log_sheet": "Log"},
            },
            "print_title_rows": "1:6",
            "page_break_enabled": False,
        }

        # --- run ---
        main(config=config)

        # --- verify ---
        out_path = out_dir / "data_test.xlsx"
        assert out_path.exists()

        wb_after = load_workbook(str(out_path))
        ws_after = wb_after["Cutsheet"]

        # Data columns C(3), D(4), E(5), G(7), H(8) — early cols, unaffected by Q/R/S deletes
        assert ws_after.cell(row=2, column=3).value == "NE1_R2", "C2 preserved"
        assert ws_after.cell(row=2, column=4).value == "PORT1_R2", "D2 preserved"
        assert ws_after.cell(row=2, column=5).value == "LBL1_R2", "E2 preserved"
        assert ws_after.cell(row=2, column=7).value == "NE2_R2", "G2 preserved"
        assert ws_after.cell(row=2, column=8).value == "PORT2_R2", "H2 preserved"

        # Row 3 and 4 also intact
        assert ws_after.cell(row=3, column=3).value == "NE1_R3"
        assert ws_after.cell(row=4, column=8).value == "PORT2_R4"

        wb_after.close()

    def test_cleanup_with_backward_compat(self, tmp_path):
        """No build_at: fall back to deleting paste_to columns."""
        # --- setup ---
        src_dir = tmp_path / "source"
        src_dir.mkdir()
        out_dir = tmp_path / "output"
        matching_path = tmp_path / "matching.xlsx"
        _make_matching(matching_path)

        src_path = src_dir / "compat.xlsx"
        _make_source(src_path, cols=19, rows=4)  # A-S

        wb_before = load_workbook(str(src_path))
        cols_before = wb_before["Cutsheet"].max_column
        wb_before.close()

        config = {
            "action": "cleanup",
            "matching_file": str(matching_path),
            "source_folder": str(src_dir),
            "target_folder": str(tmp_path / "target"),
            "output_folder": str(out_dir),
            "data_sheet": "Cutsheet",
            "target_sheet": "IP",
            "matching_sheet": "Sheet1",
            "filename_col": "Site",
            "planwork_col": "PW Number",
            "source_start_row": 2,
            "paste_start_row": 3,
            "columns": {
                "pw": {"type": "planwork", "paste_to": "Q"},
            },
            "print_title_rows": "1:6",
            "page_break_enabled": False,
        }

        # --- run ---
        main(config=config)

        # --- verify ---
        out_path = out_dir / "compat.xlsx"
        assert out_path.exists()

        wb_after = load_workbook(str(out_path))
        ws_after = wb_after["Cutsheet"]
        cols_after = ws_after.max_column

        # One column (paste_to="Q") deleted
        assert cols_after == cols_before - 1, \
            f"Expected {cols_before - 1} columns, got {cols_after}"

        # Column A data intact
        assert ws_after.cell(row=1, column=1).value == "R1C1"
        wb_after.close()

    def test_cleanup_empty_source_folder(self, tmp_path):
        """Empty source folder: main() should not raise an error."""
        # --- setup ---
        src_dir = tmp_path / "source"
        src_dir.mkdir()
        out_dir = tmp_path / "output"
        matching_path = tmp_path / "matching.xlsx"
        _make_matching(matching_path)

        config = {
            "action": "cleanup",
            "matching_file": str(matching_path),
            "source_folder": str(src_dir),
            "target_folder": str(tmp_path / "target"),
            "output_folder": str(out_dir),
            "data_sheet": "Cutsheet",
            "target_sheet": "IP",
            "matching_sheet": "Sheet1",
            "filename_col": "Site",
            "planwork_col": "PW Number",
            "source_start_row": 2,
            "paste_start_row": 3,
            "columns": {
                "pw": {"type": "planwork", "build_at": "Q", "paste_to": "J"},
            },
            "print_title_rows": "1:6",
            "page_break_enabled": False,
        }

        # --- run (must not raise) ---
        try:
            main(config=config)
        except Exception as e:
            pytest.fail(f"main() raised {type(e).__name__}: {e}")

        # output folder created but empty
        assert out_dir.exists()
