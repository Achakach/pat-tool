"""PNG inserter — sheet manipulation utilities."""

import re
import struct
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.pagebreak import Break
import math


def purge_sheet(xlsx_path: Path, sheet_name: str, from_row: int):
    """Delete all rows from *from_row* to end in the given sheet."""
    wb = load_workbook(str(xlsx_path))
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]
    if ws.max_row >= from_row:
        ws.delete_rows(from_row, ws.max_row - from_row + 1)
    wb.save(str(xlsx_path))
    wb.close()


def extract_label(filename: str) -> str:
    """Extract label from PNG filename.
    'PW planwork100_exist BKK01_Bayface Before.png' → 'Bayface Before'"""
    stem = Path(filename).stem
    stem = re.sub(r'_\d+$', '', stem)  # strip _1, _2 dedup suffix
    parts = stem.rsplit("_", 1)
    return parts[-1] if len(parts) > 1 else stem


def extract_site(filename: str) -> str:
    """Extract site name from PNG filename.
    'PW planwork100_exist BKK01_Bayface Before.png' → 'BKK01'"""
    stem = Path(filename).stem
    parts = stem.split("_", 2)  # ["PW planwork100", "exist BKK01", "Bayface Before"]
    if len(parts) >= 2:
        prefix_site = parts[1]  # "exist BKK01" or "new BKK09"
        site_parts = prefix_site.split(" ", 1)
        if len(site_parts) >= 2:
            return site_parts[1]  # "BKK01"
    return stem


def clean_sheet_name(name: str) -> str:
    """Clean sheet name for comparison.
    '2.1. Bayface_Before' → 'bayface before'"""
    name = re.sub(r'^\d+\.?\d*\.?\s*', '', name)
    name = re.sub(r'\([^)]*\)', '', name)
    name = name.replace("_", " ")
    return " ".join(name.split()).lower()


def find_matching_sheet(wb, label: str) -> str | None:
    """Find sheet whose cleaned name matches the label."""
    clean_label = clean_sheet_name(label)
    for sheet_name in wb.sheetnames:
        if clean_sheet_name(sheet_name) == clean_label:
            return sheet_name
    return None


def _setup_a4_print(ws):
    """Configure sheet for A4 portrait printing."""
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.autoPageBreaks = False
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def _calc_page_rows(ws, config_override=None):
    """Calculate max rows fitting on one A4 page based on margins.
    Paper height = 841.89 points (A4).
    Row height = 15 points (default Excel row height).
    Returns integer row count. Pass config_override to bypass auto-calc."""
    if config_override is not None:
        return config_override
    paper_height_pts = 841.89
    margins_pts = (float(ws.page_margins.top) + float(ws.page_margins.bottom)) * 72
    printable_pts = paper_height_pts - margins_pts
    return math.ceil(printable_pts / 15)


def _clear_page_breaks(ws):
    """Remove all manual horizontal page breaks from worksheet."""
    ws.row_breaks.brk = ()


def insert_png(xlsx_path: Path, sheet_name: str, png_path: Path,
               label: str, start_row: int, merge_to_col: str | None = None,
               gap_rows: int = 1, col: str = "A",
               display_width: int | None = None,
               page_rows: int | None = None, purge_from: int = 0) -> int:
    """Insert label + PNG. Returns next available row.
    When page_rows is set, inserts a page break before the label row
    (skipping the first site whose label starts at purge_from)."""
    wb = load_workbook(str(xlsx_path))
    ws = wb[sheet_name]

    # Read PNG dimensions
    with open(png_path, 'rb') as f:
        f.read(16)
        w, h = struct.unpack('>II', f.read(8))

    # Scale image to display_width (if configured)
    img = XlImage(str(png_path))
    if display_width:
        scale = display_width / w
        img.width = display_width
        img.height = int(h * scale)
        display_h = img.height
    else:
        display_h = h

    # Count rows this image needs (pixels → points → rows)
    default_ht = 15
    rows_needed = max(1, int(display_h * 0.75 / default_ht) + 1)

    # Page break before label (skip first site on sheet)
    if page_rows is not None and start_row > purge_from:
        page_end = ((start_row - 2) // page_rows + 1) * page_rows + 1
        start_row = max(start_row, page_end)
        ws.row_breaks.append(Break(id=start_row))

    # Overflow guard: if label+image group won't fit, push to next page
    if page_rows is not None:
        img_end = start_row + 1 + gap_rows + rows_needed
        page_end = ((start_row - 1) // page_rows + 1) * page_rows
        if img_end > page_end:
            start_row = page_end + 1

    # Label row (at final snapped/pushed start_row)
    label_cell = ws.cell(row=start_row, column=1)
    label_cell.value = label
    label_cell.font = Font(bold=True, size=12)
    label_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    label_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Merge label row cells (if configured)
    if merge_to_col:
        ws.merge_cells(f"A{start_row}:{merge_to_col}{start_row}")

    # Insert image (offset by gap)
    img_row = start_row + 1 + gap_rows  # label + gap + image
    ws.add_image(img, f"{col}{img_row}")

    wb.save(str(xlsx_path))
    wb.close()
    return img_row + rows_needed + gap_rows  # image + gap for next


def insert_png_no_label(xlsx_path: Path, sheet_name: str, png_path: Path,
                         start_row: int, gap_rows: int = 1, col: str = "A",
                         display_width: int | None = None,
                         page_rows: int | None = None) -> int:
    """Insert PNG without label row. Returns next available row.
    When page_rows is set, pushes the image to the next page if it
    would overflow the current page boundary."""
    wb = load_workbook(str(xlsx_path))
    ws = wb[sheet_name]
    with open(png_path, 'rb') as f:
        f.read(16)
        w, h = struct.unpack('>II', f.read(8))

    # Scale image to display_width (if configured)
    img = XlImage(str(png_path))
    if display_width:
        scale = display_width / w
        img.width = display_width
        img.height = int(h * scale)
        display_h = img.height
    else:
        display_h = h

    rows_needed = max(1, int(display_h * 0.75 / 15) + 1)

    # Page boundary overflow guard
    if page_rows is not None:
        img_end = start_row + gap_rows + rows_needed
        page_end = ((start_row - 1) // page_rows + 1) * page_rows
        if img_end > page_end:
            start_row = page_end + 1  # push image to next page

    img_row = start_row + gap_rows
    ws.add_image(img, f"{col}{img_row}")
    wb.save(str(xlsx_path))
    wb.close()
    return img_row + rows_needed + gap_rows
