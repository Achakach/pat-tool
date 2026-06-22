"""PNG inserter — sheet manipulation utilities."""

import re
import struct
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.pagebreak import Break
import math
import sys
from collections import Counter
from openpyxl.utils.units import pixels_to_points, DEFAULT_ROW_HEIGHT


def _detect_row_height(ws, fallback=None):
    """Sample row heights from existing rows. Returns MODE of explicit heights.
    Falls back to DEFAULT_ROW_HEIGHT if no explicit heights found."""
    if fallback is None:
        fallback = DEFAULT_ROW_HEIGHT
    heights = []
    for r in range(1, ws.max_row + 1):
        h = ws.row_dimensions[r].height
        if h is not None:
            heights.append(h)
    if not heights:
        return fallback
    # Use mode (most common height), not mean
    return Counter(heights).most_common(1)[0][0]


_a4_print_setup_done = False


def purge_sheet(xlsx_path: Path, sheet_name: str, from_row: int):
    """Delete all rows from *from_row* to end in the given sheet."""
    wb = load_workbook(str(xlsx_path))
    if sheet_name not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet_name]
    if ws.max_row >= from_row:
        ws.delete_rows(from_row, ws.max_row - from_row + 1)
    wb.save(str(xlsx_path))
    wb.close()


def extract_label(filename: str) -> str:
    """Extract label from PNG filename.
    'PW planwork100_exist BKK01_Bayface Before.png' → 'Bayface Before'"""
    stem = Path(filename).stem
    stem = re.sub(r'_\d+$', '', stem)  # strip _1, _2 dedup suffix
    parts = stem.rsplit("_", 1)
    return parts[-1] if len(parts) > 1 else stem


def extract_site(filename: str) -> str:
    """Extract site name from PNG filename.
    'PW planwork100_exist BKK01_Bayface Before.png' → 'BKK01'"""
    stem = Path(filename).stem
    parts = stem.split("_", 2)  # ["PW planwork100", "exist BKK01", "Bayface Before"]
    if len(parts) >= 2:
        prefix_site = parts[1]  # "exist BKK01" or "new BKK09"
        site_parts = prefix_site.split(" ", 1)
        if len(site_parts) >= 2:
            return site_parts[1]  # "BKK01"
    return stem


def clean_sheet_name(name: str) -> str:
    """Clean sheet name for comparison.
    '2.1. Bayface_Before' → 'bayface before'"""
    name = re.sub(r'^\d+\.?\d*\.?\s*', '', name)
    name = re.sub(r'\([^)]*\)', '', name)
    name = name.replace("_", " ")
    return " ".join(name.split()).lower()


def find_matching_sheet(wb, label: str) -> str | None:
    """Find sheet whose cleaned name matches the label."""
    clean_label = clean_sheet_name(label)
    for sheet_name in wb.sheetnames:
        if clean_sheet_name(sheet_name) == clean_label:
            return sheet_name
    return None


def _setup_a4_print(ws, print_title_rows=None):
    """Configure sheet for A4 portrait printing.
    
    Args:
        ws: Worksheet to configure.
        print_title_rows: Optional str like '1:6' to set as print title rows (repeat at top of every page).
    """
    global _a4_print_setup_done
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    try:
        ws.page_setup.autoPageBreaks = True
    except AttributeError:
        print("[WARNING] Could not set autoPageBreaks: worksheet XML missing pageSetupPr element", file=sys.stderr)
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    if print_title_rows is not None:
        ws.print_title_rows = print_title_rows
    if not _a4_print_setup_done:
        print(f"[DEBUG] _setup_a4_print: setting A4 portrait, margins L=0.25 R=0.25 T=0.5 B=0.5, autoPageBreaks=True, print_title_rows={print_title_rows}", file=sys.stderr)
        _a4_print_setup_done = True


def _calc_page_rows(ws, config_override=None):
    """Calculate max rows fitting on one A4 page based on margins.
    Paper height = 841.89 points (A4).
    Row height = detected from worksheet row_dimensions.
    Returns integer row count. Pass config_override to bypass auto-calc."""
    if config_override is not None:
        print(f"[DEBUG] _calc_page_rows: config_override={config_override}, returning early", file=sys.stderr)
        return config_override
    paper_height_pts = 841.89
    margins_pts = (float(ws.page_margins.top) + float(ws.page_margins.bottom)) * 72
    printable_pts = paper_height_pts - margins_pts
    row_height = _detect_row_height(ws)
    rows = math.ceil(printable_pts / row_height)
    print(f"[DEBUG] _calc_page_rows: margins top={ws.page_margins.top} bottom={ws.page_margins.bottom}, paper=841.89, margins_pts={margins_pts}, printable_pts={printable_pts}, row_height={row_height}, rows={printable_pts}/{row_height}={rows} (ceil)", file=sys.stderr)
    return rows


def _clear_page_breaks(ws):
    """Remove all manual horizontal page breaks from worksheet."""
    ws.row_breaks.brk = ()


def insert_png(xlsx_path: Path, sheet_name: str, png_path: Path,
               label: str, start_row: int, merge_to_col: str | None = None,
               gap_rows: int = 1, col: str = "A",
               display_width: int | None = None,
               page_rows: int | None = None, purge_from: int = 0,
               header_count: int = 0) -> int:
    """Insert label + PNG. Returns next available row.
    When page_rows is set, inserts a page break before the label row
    (skipping the first site whose label starts at purge_from)."""
    wb = load_workbook(str(xlsx_path))
    ws = wb[sheet_name]
    _sr0 = start_row  # capture original for debug
    print(f"[DEBUG] insert_png: sheet='{sheet_name}' site='{label}' start_row={_sr0} purge_from={purge_from} page_rows={page_rows} header_count={header_count}", file=sys.stderr)

    # Read PNG dimensions
    with open(png_path, 'rb') as f:
        f.read(16)
        w, h = struct.unpack('>II', f.read(8))

    # Scale image to display_width (if configured)
    img = XlImage(str(png_path))
    if display_width:
        scale = display_width / w
        img.width = display_width
        img.height = int(h * scale)
        display_h = img.height
    else:
        display_h = h

    # Count rows this image needs (pixels → points → rows)
    row_ht = _detect_row_height(ws)
    height_pts = pixels_to_points(display_h)
    rows_needed = max(1, math.ceil(height_pts / row_ht))

    # Snap to next page boundary (skip first site on sheet)
    if page_rows is not None and start_row > purge_from:
        if header_count and header_count > 0:
            content_rows = page_rows - header_count
            if start_row <= page_rows:
                page_end = page_rows + 1  # snap to page 2 start
            else:
                offset = start_row - page_rows - 2  # -2 preserves exact page boundaries
                pages_after = offset // content_rows + 1
                page_end = page_rows + 1 + pages_after * content_rows
            print(f"[DEBUG] insert_png:   snap: header_count={header_count} content_rows={content_rows} page_end={page_end}", file=sys.stderr)
        else:
            page_end = ((start_row - 2) // page_rows + 1) * page_rows + 1
            print(f"[DEBUG] insert_png:   snap: page_end = (({start_row}-2)//{page_rows}+1)*{page_rows}+1 = {page_end}", file=sys.stderr)
        new_sr = max(start_row, page_end)
        print(f"[DEBUG] insert_png:   snap check: start_row({start_row}) > purge_from({purge_from})? YES, start_row = max({start_row},{page_end}) = {new_sr}", file=sys.stderr)
        start_row = new_sr
    elif page_rows is not None:
        print(f"[DEBUG] insert_png:   snap check: start_row({start_row}) > purge_from({purge_from})? NO (page_rows={page_rows})", file=sys.stderr)

    # Overflow guard: if label+image group won't fit, push to next page
    if page_rows is not None:
        img_end = start_row + 1 + gap_rows + rows_needed
        if header_count and header_count > 0:
            content_rows = page_rows - header_count
            if start_row <= page_rows:
                page_end = page_rows
            else:
                offset = start_row - page_rows - 1
                pages_before = offset // content_rows
                page_end = page_rows + (pages_before + 1) * content_rows
        else:
            page_end = ((start_row - 1) // page_rows + 1) * page_rows
        overflow = img_end > page_end
        if header_count and header_count > 0:
            print(f"[DEBUG] insert_png:   overflow: rows_needed={rows_needed}, img_end={start_row}+1+{gap_rows}+{rows_needed}={img_end}, page_end={page_end}, header_count={header_count}, content_rows={content_rows}, {img_end}>{page_end}? {'YES' if overflow else 'NO'}", file=sys.stderr)
        else:
            print(f"[DEBUG] insert_png:   overflow: rows_needed={rows_needed}, img_end={start_row}+1+{gap_rows}+{rows_needed}={img_end}, page_end={page_end}, {img_end}>{page_end}? {'YES' if overflow else 'NO'}", file=sys.stderr)
        if overflow:
            start_row = page_end + 1
            print(f"[DEBUG] insert_png:   pushed to row {start_row}", file=sys.stderr)

    # Label row (at final snapped/pushed start_row)
    label_cell = ws.cell(row=start_row, column=1)
    label_cell.value = label
    label_cell.font = Font(bold=True, size=12)
    label_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    label_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Merge label row cells (if configured)
    if merge_to_col:
        ws.merge_cells(f"A{start_row}:{merge_to_col}{start_row}")

    # Insert image (offset by gap)
    img_row = start_row + 1 + gap_rows  # label + gap + image
    ws.add_image(img, f"{col}{img_row}")

    wb.save(str(xlsx_path))
    wb.close()
    next_row = img_row + rows_needed + gap_rows  # image + gap for next
    print(f"[DEBUG] insert_png:   FINAL: label at row {start_row}, image at row {img_row}, returns {next_row}", file=sys.stderr)
    return next_row


def insert_png_no_label(xlsx_path: Path, sheet_name: str, png_path: Path,
                         start_row: int, gap_rows: int = 1, col: str = "A",
                         display_width: int | None = None,
                         page_rows: int | None = None,
                         header_count: int = 0) -> int:
    """Insert PNG without label row. Returns next available row.
    When page_rows is set, pushes the image to the next page if it
    would overflow the current page boundary."""
    wb = load_workbook(str(xlsx_path))
    ws = wb[sheet_name]
    print(f"[DEBUG] insert_png_no_label: sheet='{sheet_name}' start_row={start_row} page_rows={page_rows}", file=sys.stderr)
    with open(png_path, 'rb') as f:
        f.read(16)
        w, h = struct.unpack('>II', f.read(8))

    # Scale image to display_width (if configured)
    img = XlImage(str(png_path))
    if display_width:
        scale = display_width / w
        img.width = display_width
        img.height = int(h * scale)
        display_h = img.height
    else:
        display_h = h

    row_ht = _detect_row_height(ws)
    height_pts = pixels_to_points(display_h)
    rows_needed = max(1, math.ceil(height_pts / row_ht))

    # Page boundary overflow guard (header_count-aware)
    if page_rows is not None:
        img_end = start_row + gap_rows + rows_needed
        if header_count and header_count > 0:
            content_rows = page_rows - header_count
            if start_row <= page_rows:
                page_end = page_rows
            else:
                offset = start_row - page_rows - 1
                pages_before = offset // content_rows
                page_end = page_rows + (pages_before + 1) * content_rows
            overflow = img_end > page_end
            print(f"[DEBUG] insert_png_no_label:   rows_needed={rows_needed}, header_count={header_count}, content_rows={content_rows}, img_end={start_row}+{gap_rows}+{rows_needed}={img_end}, page_end={page_end}, {img_end}>{page_end}? {'YES' if overflow else 'NO'}", file=sys.stderr)
        else:
            page_end = ((start_row - 1) // page_rows + 1) * page_rows
            overflow = img_end > page_end
            print(f"[DEBUG] insert_png_no_label:   rows_needed={rows_needed}, img_end={start_row}+{gap_rows}+{rows_needed}={img_end}, page_end=(({start_row}-1)//{page_rows}+1)*{page_rows}={page_end}, {img_end}>{page_end}? {'YES' if overflow else 'NO'}", file=sys.stderr)
        if overflow:
            start_row = page_end + 1
            print(f"[DEBUG] insert_png_no_label:   pushed to row {start_row}", file=sys.stderr)

    img_row = start_row + gap_rows
    ws.add_image(img, f"{col}{img_row}")
    wb.save(str(xlsx_path))
    wb.close()
    next_row = img_row + rows_needed + gap_rows
    print(f"[DEBUG] insert_png_no_label:   FINAL: image at row {img_row}, returns {next_row}", file=sys.stderr)
    return next_row
