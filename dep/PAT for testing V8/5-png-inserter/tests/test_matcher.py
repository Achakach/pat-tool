"""Tests for png-inserter matcher and inserter."""
import sys
from pathlib import Path
import pytest
from openpyxl import Workbook, load_workbook
from src.matcher import read_matching, match_pngs, extract_planwork
from src.inserter import extract_label, extract_site, clean_sheet_name, find_matching_sheet, insert_png


class TestReadMatching:
    def test_basic_mapping(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "filename"   # header
        ws["B1"] = "planwork"   # header
        ws["A2"] = "one.xlsx"
        ws["B2"] = "siteA"
        ws["A3"] = "two.xlsx"
        ws["B3"] = "siteC"
        path = tmp_path / "match.xlsx"
        wb.save(str(path))
        wb.close()

        result = read_matching(str(path), "Sheet", "filename", "planwork")
        assert result == {"one.xlsx": ["siteA"], "two.xlsx": ["siteC"]}

    def test_inherit_blank_filename(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "filename"   # header
        ws["B1"] = "planwork"   # header
        ws["A2"] = "one.xlsx"
        ws["B2"] = "siteA"
        ws["A3"] = None         # blank → inherit "one.xlsx"
        ws["B3"] = "siteB"
        path = tmp_path / "match.xlsx"
        wb.save(str(path))
        wb.close()

        result = read_matching(str(path), "Sheet", "filename", "planwork")
        assert result == {"one.xlsx": ["siteA", "siteB"]}

    def test_skip_rows_before_first_filename(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "filename"   # header
        ws["B1"] = "planwork"   # header
        ws["B2"] = "siteA"      # no filename in row 2 → skipped
        ws["A3"] = "one.xlsx"
        ws["B3"] = "siteB"
        path = tmp_path / "match.xlsx"
        wb.save(str(path))
        wb.close()

        result = read_matching(str(path), "Sheet", "filename", "planwork")
        assert "siteA" not in str(result)


class TestMatchPngs:
    def test_match_by_planwork(self, tmp_path):
        (tmp_path / "PW planwork100_exist BKK01_Bayface Before.png").touch()
        (tmp_path / "PW test001_new BKK09_Bayface Before.png").touch()
        (tmp_path / "unrelated.png").touch()

        matches = match_pngs(tmp_path, ["planwork100"])
        names = [p.name for p in matches]
        assert "PW planwork100_exist BKK01_Bayface Before.png" in names
        assert "unrelated.png" not in names

    def test_multiple_planworks(self, tmp_path):
        (tmp_path / "PW planwork100_exist BKK01_Bayface.png").touch()
        (tmp_path / "PW test001_new BKK09_Bayface.png").touch()

        matches = match_pngs(tmp_path, ["planwork100", "test001"])
        assert len(matches) == 2

    def test_no_matches(self, tmp_path):
        (tmp_path / "PW planwork100_exist BKK01_Bayface.png").touch()
        matches = match_pngs(tmp_path, ["nonexistent"])
        assert len(matches) == 0


class TestPurgeSheet:
    def test_purge_sheet(self, tmp_path):
        from openpyxl import Workbook
        from src.inserter import purge_sheet

        wb = Workbook()
        ws = wb.active
        for i in range(1, 20):
            ws.cell(row=i, column=1).value = f"row{i}"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        wb.close()

        purge_sheet(path, "Sheet", 10)

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2.max_row == 9  # rows 10-19 deleted
        assert ws2.cell(row=9, column=1).value == "row9"
        wb2.close()

    def test_purge_from_row_beyond_max(self, tmp_path):
        """Purge from row > max_row should be a no-op."""
        from openpyxl import Workbook
        from src.inserter import purge_sheet

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "keep"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        wb.close()

        purge_sheet(path, "Sheet", 99)

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2.max_row >= 1
        assert ws2["A1"].value == "keep"
        wb2.close()

    def test_purge_nonexistent_sheet(self, tmp_path):
        """Purge on a sheet that doesn't exist should be a no-op."""
        from openpyxl import Workbook
        from src.inserter import purge_sheet

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "data"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        wb.close()

        purge_sheet(path, "NoSuchSheet", 5)

        wb2 = load_workbook(str(path))
        ws2 = wb2.active
        assert ws2["A1"].value == "data"
        wb2.close()


class TestSheetMatching:
    def test_extract_label(self):
        assert extract_label("PW planwork100_exist BKK01_Bayface Before.png") == "Bayface Before"
        assert extract_label("PW xxx_new siteB_Alarm After.png") == "Alarm After"
        assert extract_label("PW planwork100_exist BKK01_Bayface Before_1.png") == "Bayface Before"

    def test_extract_site(self):
        assert extract_site("PW planwork100_exist BKK01_Bayface Before.png") == "BKK01"
        assert extract_site("PW xxx_new BKK09_Alarm After.png") == "BKK09"
        assert extract_site("PW planwork_exist BKK10_Summary.png") == "BKK10"

    def test_clean_sheet_name(self):
        assert clean_sheet_name("2.1. Bayface_Before") == "bayface before"
        assert clean_sheet_name("4.1 Alarm Before(7)") == "alarm before"
        assert clean_sheet_name("Summary") == "summary"

    def test_find_matching_sheet(self, tmp_path):
        wb = Workbook()
        wb.active.title = "2.1. Bayface_Before"
        wb.create_sheet("4.1 Alarm Before(7)")
        wb.create_sheet("Summary")
        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        wb.close()

        wb2 = load_workbook(str(path))
        assert find_matching_sheet(wb2, "Bayface Before") == "2.1. Bayface_Before"
        assert find_matching_sheet(wb2, "Alarm Before") == "4.1 Alarm Before(7)"
        assert find_matching_sheet(wb2, "Summary") == "Summary"
        assert find_matching_sheet(wb2, "No Match") is None
        wb2.close()


class TestInsertPng:
    def test_insert_creates_label_and_image(self, tmp_path):
        """Insert label + PNG, verify label cell and next-row return."""
        import struct
        import zlib

        def create_png(w, h, color=(255, 0, 0)):
            def chunk(typ, data):
                c = typ + data
                return struct.pack('>I', len(data)) + c + struct.pack('>I', zlib.crc32(c) & 0xffffffff)
            ihdr = struct.pack('>IIBBBBB', w, h, 8, 2, 0, 0, 0)
            raw = b''
            for _ in range(h):
                raw += b'\x00'
                for _ in range(w):
                    raw += bytes(color)
            return b'\x89PNG\r\n\x1a\n' + chunk(b'IHDR', ihdr) + chunk(b'IDAT', zlib.compress(raw)) + chunk(b'IEND', b'')

        png_path = tmp_path / "test.png"
        png_path.write_bytes(create_png(10, 10, (100, 150, 200)))

        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        next_row = insert_png(xlsx_path, "Test", png_path, "BKK01", 10)
        assert next_row == 14  # 10(label) + 1(gap) + 1(img) + 1(gap) + 1(img offset) = 14

        wb2 = load_workbook(str(xlsx_path))
        ws2 = wb2["Test"]
        assert ws2.cell(row=10, column=1).value == "BKK01"
        assert ws2.cell(row=10, column=1).font.bold is True
        wb2.close()


class TestMainIntegration:
    """Integration tests for the full insert.py pipeline via config injection."""

    def test_main_with_config_dict(self, tmp_path):
        """Full integration: matching + PNG + XLSX → output with image inserted."""
        # Minimal valid 1x1 PNG bytes
        _PNG_DATA = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
            b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00"
            b"\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82"
        )

        # ── Setup directories ──────────────────────────────────────────
        match_dir = tmp_path
        png_dir = tmp_path / "pngs"
        xlsx_dir = tmp_path / "xlsx"
        out_dir = tmp_path / "output"
        png_dir.mkdir()
        xlsx_dir.mkdir()

        # ── Create matching.xlsx ───────────────────────────────────────
        match_wb = Workbook()
        match_ws = match_wb.active
        match_ws["A1"] = "Site"
        match_ws["B1"] = "PW Number"
        match_ws["A2"] = "TestSite"
        match_ws["B2"] = "XX001"
        match_path = match_dir / "matching.xlsx"
        match_wb.save(str(match_path))
        match_wb.close()

        # ── Create PNG file ────────────────────────────────────────────
        png_path = png_dir / "PW XX001_exist TestSite_label.png"
        png_path.write_bytes(_PNG_DATA)

        # ── Create XLSX with matching sheet ────────────────────────────
        xlsx_wb = Workbook()
        xlsx_ws = xlsx_wb.active
        xlsx_ws.title = "label"
        # Fill enough rows so purge_from_row=5 has something to delete
        for i in range(1, 12):
            xlsx_ws.cell(row=i, column=1).value = f"row{i}"
        xlsx_path = xlsx_dir / "TestSite.xlsx"
        xlsx_wb.save(str(xlsx_path))
        xlsx_wb.close()

        # ── Run main with config dict injection ────────────────────────
        config = {
            "matching_file": str(match_path),
            "xlsx_folder": str(xlsx_dir),
            "png_folder": str(png_dir),
            "output_folder": str(out_dir),
            "matching_sheet": "Sheet",
            "filename_col": "Site",
            "planwork_col": "PW Number",
            "purge_from_row": 5,
            "label_merge_to_col": "G",
            "insert_gap_rows": 1,
            "image_insert_col": "A",
            "print_title_rows": None,
            "page_break_before_label": False,
        }
        from insert import main
        main(config=config)

        # ── Verify output ──────────────────────────────────────────────
        output_xlsx = out_dir / "TestSite.xlsx"
        assert output_xlsx.exists(), (
            f"Output XLSX not found at {output_xlsx}"
        )

        # Verify it's a valid workbook
        wb = load_workbook(str(output_xlsx))
        assert "label" in wb.sheetnames, (
            f"Sheet 'label' not found in output. Sheets: {wb.sheetnames}"
        )
        ws = wb["label"]
        # Row 5 should have the site label "TestSite" (bold)
        label_cell = ws.cell(row=5, column=1)
        assert label_cell.value == "TestSite", (
            f"Expected 'TestSite' at A5, got '{label_cell.value}'"
        )
        assert label_cell.font.bold is True, "Label should be bold"
        wb.close()

