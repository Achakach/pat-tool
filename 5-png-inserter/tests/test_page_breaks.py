"""Tests for page break logic in png-inserter — Tasks 4, 5, 6."""
import struct
import zlib
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.pagebreak import Break

from src.inserter import (
    _calc_page_rows,
    _clear_page_breaks,
    _detect_row_height,
    _setup_a4_print,
    insert_png,
    insert_png_no_label,
)


def _make_test_png(path, width=10, height=10):
    """Create a minimal valid PNG for testing."""
    def _chunk(chunk_type, data):
        c = chunk_type + data
        return struct.pack('>I', len(data)) + c + struct.pack('>I', zlib.crc32(c) & 0xffffffff)

    ihdr = struct.pack('>IIBBBBB', width, height, 8, 2, 0, 0, 0)
    raw = b''
    for _ in range(height):
        raw += b'\x00'  # filter byte
        for _ in range(width):
            raw += b'\xff\x00\x00'  # RGB red pixels
    with open(path, 'wb') as f:
        f.write(b'\x89PNG\r\n\x1a\n')
        f.write(_chunk(b'IHDR', ihdr))
        f.write(_chunk(b'IDAT', zlib.compress(raw)))
        f.write(_chunk(b'IEND', b''))


def _make_test_xlsx(tmp_path, name="test.xlsx"):
    """Create an xlsx with Sheet as active sheet and some data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for i in range(1, 20):
        ws.cell(row=i, column=1).value = f"row{i}"
    path = tmp_path / name
    wb.save(str(path))
    wb.close()
    return path


# ============================================================
# Task 1 — _detect_row_height Tests (4 tests)
# ============================================================

class TestDetectRowHeight:
    """Tests for _detect_row_height helper function."""

    def test_all_explicit_20pt_returns_20(self):
        """All rows at 20pt → mode = 20.0."""
        wb = Workbook()
        ws = wb.active
        for r in range(1, 11):
            ws.cell(row=r, column=1).value = f"row{r}"
            ws.row_dimensions[r].height = 20.0
        assert _detect_row_height(ws) == 20.0
        wb.close()

    def test_mixed_returns_mode_15(self):
        """5 rows at 24pt + 10 rows at 15pt → mode = 15.0."""
        wb = Workbook()
        ws = wb.active
        for r in range(1, 6):
            ws.cell(row=r, column=1).value = f"header{r}"
            ws.row_dimensions[r].height = 24.0
        for r in range(6, 16):
            ws.cell(row=r, column=1).value = f"content{r}"
            ws.row_dimensions[r].height = 15.0
        assert _detect_row_height(ws) == 15.0
        wb.close()

    def test_no_explicit_returns_fallback(self):
        """No explicit heights set → fallback to DEFAULT_ROW_HEIGHT (15.0)."""
        wb = Workbook()
        ws = wb.active
        for r in range(1, 11):
            ws.cell(row=r, column=1).value = f"row{r}"
            # Do NOT set row_dimensions height — stays None
        assert _detect_row_height(ws) == 15.0
        wb.close()

    def test_empty_worksheet_returns_fallback(self):
        """Empty worksheet (no rows) → fallback to DEFAULT_ROW_HEIGHT (15.0)."""
        wb = Workbook()
        ws = wb.active
        # No data rows at all — max_row should be 0 or 1 with no data
        assert _detect_row_height(ws) == 15.0
        wb.close()


# ============================================================
# Task 2 — Print Title Rows Tests (2 tests)
# ============================================================

class TestSetupA4PrintTitleRows:
    """Tests for _setup_a4_print print_title_rows parameter."""

    def test_setup_a4_print_sets_print_title_rows(self, tmp_path):
        """_setup_a4_print(ws, '1:6') sets ws.print_title_rows to '1:6'."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws, "1:6")
        assert ws.print_title_rows == "$1:$6"
        wb.close()

    def test_setup_a4_print_none_does_not_set(self, tmp_path):
        """_setup_a4_print(ws, None) leaves ws.print_title_rows unset."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws, None)
        assert ws.print_title_rows is None
        wb.close()


# ============================================================
# Task 4 — Config / Setup Tests (7 tests)
# ============================================================

class TestPageBreakConfig:
    """Tests for page break feature toggle and A4 setup helpers."""

    def test_page_break_before_label_true_enables_feature(self, tmp_path):
        """_calc_page_rows returns non-None int when feature conceptually enabled."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws)
        result = _calc_page_rows(ws)
        # Feature enabled: page_rows is an integer (not None)
        assert result is not None
        assert isinstance(result, int)
        assert result > 0
        wb.close()

    def test_page_break_before_label_false_disables_feature(self, tmp_path):
        """page_rows=None passed to insert_png → no breaks inserted."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "test.png"
        _make_test_png(png, 10, 10)

        insert_png(output, "Sheet", png, "Site1", 5, page_rows=None, purge_from=1, gap_rows=1)

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        assert len(ws.row_breaks.brk) == 0
        wb.close()

    def test_page_break_before_label_missing_defaults_false(self, tmp_path):
        """page_rows not passed (default None) → no breaks (backward compat)."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "test.png"
        _make_test_png(png, 10, 10)

        insert_png(output, "Sheet", png, "Site1", 5, purge_from=1, gap_rows=1)

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        assert len(ws.row_breaks.brk) == 0
        wb.close()

    def test_a4_page_rows_override(self, tmp_path):
        """Config a4_page_rows=40 → _calc_page_rows returns 40."""
        wb = Workbook()
        ws = wb.active
        result = _calc_page_rows(ws, config_override=40)
        assert result == 40
        wb.close()

    def test_calc_page_rows_20pt(self):
        """20pt explicit rows → ceil(769.89 / 20) = 39."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws)
        for r in range(1, 11):
            ws.cell(row=r, column=1).value = f"row{r}"
            ws.row_dimensions[r].height = 20.0
        result = _calc_page_rows(ws)
        assert result == 39  # ceil(769.89 / 20)
        wb.close()

    def test_calc_page_rows_mixed_mode(self):
        """Mixed heights 24pt (5 rows) + 15pt (10 rows) → mode 15 → ceil(769.89 / 15) = 52."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws)
        for r in range(1, 6):
            ws.cell(row=r, column=1).value = f"header{r}"
            ws.row_dimensions[r].height = 24.0
        for r in range(6, 16):
            ws.cell(row=r, column=1).value = f"content{r}"
            ws.row_dimensions[r].height = 15.0
        result = _calc_page_rows(ws)
        assert result == 52  # mode is 15 (10 rows) > ceil(769.89/15) = 52
        wb.close()

    def test_a4_page_rows_absent_autocalc(self, tmp_path):
        """No a4_page_rows → _calc_page_rows returns ~51 (auto-calc)."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws)
        result = _calc_page_rows(ws)
        # A4 auto-calc with 0.5" top+bottom margins → (841.89 - 72) / 15 ≈ 52
        assert result == 52
        wb.close()

    def test_auto_page_breaks_enabled(self, tmp_path):
        """_setup_a4_print → autoPageBreaks is True."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws)
        assert ws.page_setup.autoPageBreaks is True
        wb.close()

    def test_no_crash_on_missing_page_setup_pr(self, tmp_path):
        """_setup_a4_print should not crash when worksheet XML lacks pageSetupPr."""
        fixture = Path(__file__).parent / "fixtures" / "no_page_setup_pr.xlsx"
        wb = load_workbook(str(fixture))
        ws = wb.active
        try:
            _setup_a4_print(ws)
            passed = True
        except AttributeError:
            passed = False
        wb.close()
        assert passed, "_setup_a4_print crashed with AttributeError on fixture"

    def test_clear_page_breaks_empties_brk(self, tmp_path):
        """_clear_page_breaks clears all existing manual breaks."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        ws.row_breaks.append(Break(id=10))
        ws.row_breaks.append(Break(id=25))
        assert len(ws.row_breaks.brk) == 2

        _clear_page_breaks(ws)
        assert len(ws.row_breaks.brk) == 0
        wb.close()


# ============================================================
# Task 5 — Break Insertion Logic Tests (6 tests)
# ============================================================

class TestPageBreakInsertion:
    """Tests for page break insertion during insert_png and insert_png_no_label."""

    def test_snap_to_page_boundary(self, tmp_path):
        """Two sites on same sheet → second site snaps to page boundary."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "test.png"
        _make_test_png(png, 10, 10)  # rows_needed = 1

        # First site at purge_from (5): no snap
        r1 = insert_png(output, "Sheet", png, "Site1", 5, page_rows=10, purge_from=5, gap_rows=1)
        # start_row=5, start_row > purge_from? No → no snap
        # label=5, img=7, rows_needed=1 → returns 7+1+1 = 9

        # Second site at row 9: should snap to page boundary (auto breaks handle pagination)
        r2 = insert_png(output, "Sheet", png, "Site2", r1, page_rows=10, purge_from=5, gap_rows=1)
        # start_row=9 > 5 → snap: ((9-2)//10+1)*10+1 = (0+1)*10+1 = 11
        # Label at 11, img at 13, returns 15

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        # No manual breaks inserted — autoPageBreaks=True handles pagination
        assert len(ws.row_breaks.brk) == 0
        assert ws.cell(row=11, column=1).value == "Site2"
        wb.close()

    def test_no_break_before_first_site(self, tmp_path):
        """First site at purge_from row → no page break inserted."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "test.png"
        _make_test_png(png, 10, 10)

        insert_png(output, "Sheet", png, "Site1", 5, page_rows=10, purge_from=5, gap_rows=1)
        # start_row=5, purge_from=5 → start_row > purge_from? No → skip break

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        assert len(ws.row_breaks.brk) == 0
        assert ws.cell(row=5, column=1).value == "Site1"
        wb.close()

    def test_overflow_guard_pushes_image(self, tmp_path):
        """Image near page boundary → insert_png_no_label pushes to next page."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "tall.png"
        _make_test_png(png, 10, 100)  # rows_needed = ceil(75/15) = 5

        # start_row=4, page_rows=5, gap_rows=1
        # img_end = 4+1+5 = 10, page_end = ((4-1)//5+1)*5 = 5
        # 10 > 5 → pushed to page_end+1 = 6
        next_row = insert_png_no_label(output, "Sheet", png, start_row=4, gap_rows=1, page_rows=5)
        # img_row = 6+1 = 7, return = 7+5+1 = 13
        assert next_row > 5  # image pushed past page boundary
        assert next_row == 13

    def test_image_fits_no_push(self, tmp_path):
        """Image fits within page boundary → stays at original start_row."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "small.png"
        _make_test_png(png, 10, 10)  # rows_needed = 1

        # start_row=4, page_rows=10, gap_rows=1
        # img_end = 4+1+1 = 6, page_end = ((4-1)//10+1)*10 = 10
        # 6 <= 10 → no push, stays at row 4
        next_row = insert_png_no_label(output, "Sheet", png, start_row=4, gap_rows=1, page_rows=10)
        # img_row = 4+1 = 5, return = 5+1+1 = 7
        assert next_row == 7  # no push, image at row 5

    def test_multiple_small_images_fill_page(self, tmp_path):
        """Third of three images overflows page → pushed to next page."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "med.png"
        _make_test_png(png, 10, 60)  # rows_needed = ceil(45/15) = 3

        # page_rows=15, gap_rows=1
        r1 = insert_png_no_label(output, "Sheet", png, start_row=1, gap_rows=1, page_rows=15)
        # img_end = 1+1+3=5, page_end=15 → stays. img_row=2, return=2+3+1=6
        r2 = insert_png_no_label(output, "Sheet", png, start_row=r1, gap_rows=1, page_rows=15)
        # img_end = 6+1+3=10, page_end=15 → stays. img_row=7, return=7+3+1=11
        r3 = insert_png_no_label(output, "Sheet", png, start_row=r2, gap_rows=1, page_rows=15)
        # img_end = 11+1+3=15, page_end=15. 15>15? NO → stays
        # img_row=12, return=12+3+1=16
        assert r1 == 6
        assert r2 == 11
        assert r3 == 16

    def test_no_manual_breaks_with_auto(self, tmp_path):
        """Auto page breaks enabled → no manual breaks inserted."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "test.png"
        _make_test_png(png, 10, 10)

        insert_png(output, "Sheet", png, "Site1", 5, page_rows=10, purge_from=1, gap_rows=1)
        insert_png(output, "Sheet", png, "Site2", 15, page_rows=10, purge_from=1, gap_rows=1)

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        breaks = ws.row_breaks.brk
        assert len(breaks) == 0
        wb.close()


# ============================================================
# Task 6 — Edge Case Tests (4 tests)
# ============================================================

class TestPageBreakEdgeCases:
    """Edge case tests for page break behavior."""

    def test_single_image_taller_than_page(self, tmp_path):
        """Image taller than page_rows → still inserted without crash (warning in orchestrator)."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "huge.png"
        _make_test_png(png, 10, 500)  # rows_needed = ceil(375/15) = 25

        # page_rows=5, image spans ~25 rows → much taller than a page
        next_row = insert_png(output, "Sheet", png, "Site1", 3, page_rows=5, purge_from=1, gap_rows=1)
        # start_row=3 > 1 → snap: ((3-2)//5+1)*5+1 = 6, snap to row 6
        # overflow guard: img_end = 6+1+1+25 = 33 > page_end=10 → push to row 11
        # label at 11, img at 13, return = 13+25+1 = 39
        assert next_row > 6  # image inserted past the page boundary

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        assert ws.cell(row=11, column=1).value == "Site1"  # label at overflow-pushed row
        wb.close()

    def test_multi_sheet_no_manual_breaks(self, tmp_path):
        """Auto breaks on one sheet do not affect another — no manual breaks anywhere."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")

        for i in range(1, 20):
            ws1.cell(row=i, column=1).value = f"s1_{i}"
            ws2.cell(row=i, column=1).value = f"s2_{i}"

        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        wb.close()

        png = tmp_path / "test.png"
        _make_test_png(png, 10, 10)

        # Sheet1: insert with purge_from=1, start_row=10 → 10>1 triggers snap
        insert_png(path, "Sheet1", png, "Site1", 10, page_rows=5, purge_from=1, gap_rows=1)

        # Sheet2: insert with purge_from=5, start_row=5 → 5>5 is False → no snap (first site)
        insert_png(path, "Sheet2", png, "SiteA", 5, page_rows=5, purge_from=5, gap_rows=1)

        wb = load_workbook(str(path))
        breaks1 = wb["Sheet1"].row_breaks.brk
        breaks2 = wb["Sheet2"].row_breaks.brk

        # No manual breaks — autoPageBreaks=True handles pagination
        assert len(breaks1) == 0
        assert len(breaks2) == 0
        wb.close()

    def test_gap_rows_zero(self, tmp_path):
        """Gap_rows=0 does not break overflow math."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "med.png"
        _make_test_png(png, 10, 60)  # rows_needed = 3

        # page_rows=5, gap_rows=0, start_row=3
        # img_end = 3+0+3 = 6, page_end = ((3-1)//5+1)*5 = 5
        # 6 > 5 → pushed: start_row = 6
        # img_row = 6+0 = 6, return = 6+3+0 = 9
        next_row = insert_png_no_label(output, "Sheet", png, start_row=3, gap_rows=0, page_rows=5)
        assert next_row >= 9  # pushed past page boundary

        # Verify image was actually inserted (no crash)
        wb = load_workbook(str(output))
        images = wb["Sheet"]._images
        assert len(images) == 1
        wb.close()

    def test_existing_breaks_cleared_when_disabled(self, tmp_path):
        """_clear_page_breaks removes pre-existing manual page breaks."""
        wb = Workbook()
        ws = wb.active
        _setup_a4_print(ws)

        # Simulate pre-existing breaks from a previous run
        ws.row_breaks.append(Break(id=15))
        ws.row_breaks.append(Break(id=30))
        ws.row_breaks.append(Break(id=45))
        assert len(ws.row_breaks.brk) == 3

        _clear_page_breaks(ws)
        assert len(ws.row_breaks.brk) == 0
        wb.close()


# ============================================================
# Task 1 — Print Title Rows Config Parsing Tests (5 tests)
# ============================================================

class TestPrintTitleRows:
    """Tests for _parse_print_title_rows config parsing."""

    def test_parse_print_title_rows_valid(self):
        """"1:6" → (6, "1:6") — standard valid input."""
        from insert import _parse_print_title_rows
        result = _parse_print_title_rows("1:6")
        assert result == (6, "1:6")

    def test_parse_print_title_rows_null(self):
        """None → (0, None) — disabled state."""
        from insert import _parse_print_title_rows
        result = _parse_print_title_rows(None)
        assert result == (0, None)

    def test_parse_print_title_rows_malformed_colon_only(self, capsys):
        """"1:" → (0, None) + stderr warning."""
        from insert import _parse_print_title_rows
        result = _parse_print_title_rows("1:")
        assert result == (0, None)
        stderr = capsys.readouterr().err
        assert "WARNING" in stderr

    def test_parse_print_title_rows_malformed_non_numeric(self, capsys):
        """"a:b" → (0, None) + stderr warning."""
        from insert import _parse_print_title_rows
        result = _parse_print_title_rows("a:b")
        assert result == (0, None)
        stderr = capsys.readouterr().err
        assert "WARNING" in stderr

    def test_parse_print_title_rows_degenerate(self, capsys):
        """"1:51" with page_rows=52 → (51, "1:51") + stderr warning (content_rows=1)."""
        from insert import _parse_print_title_rows
        result = _parse_print_title_rows("1:51", page_rows=52)
        assert result == (51, "1:51")
        stderr = capsys.readouterr().err
        assert "WARNING" in stderr
        assert "1" in stderr  # content_rows=1 mentioned in warning

    def test_overflow_with_headers_pushes(self, tmp_path):
        """page_rows=10, header_count=2, start_row=9 → image pushed past row 10."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "tall.png"
        _make_test_png(png, 10, 100)  # rows_needed = ceil(75/15) = 5

        next_row = insert_png(output, "Sheet", png, "Site1", start_row=9,
                              page_rows=10, header_count=2, purge_from=1, gap_rows=0)

        assert next_row == 17  # label@11, img@12, return=12+5+0=17

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        assert ws.cell(row=11, column=1).value == "Site1"  # label at snapped/pushed row
        wb.close()

    def test_overflow_no_headers_unchanged(self, tmp_path):
        """Same scenario with header_count=0 → original overflow behavior preserved."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "tall.png"
        _make_test_png(png, 10, 100)  # rows_needed = ceil(75/15) = 5

        next_row = insert_png(output, "Sheet", png, "Site1", start_row=9,
                              page_rows=10, header_count=0, purge_from=1, gap_rows=0)

        assert next_row == 17  # label@11, img@12, return=12+5+0=17

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        assert ws.cell(row=11, column=1).value == "Site1"
        wb.close()

    def test_snap_with_headers_keeps_boundary(self, tmp_path):
        """start_row=53, page_rows=52, header_count=6, purge_from=10 → label at row 53 (boundary preserved)."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "test.png"
        _make_test_png(png, 10, 10)  # rows_needed = 1

        next_row = insert_png(output, "Sheet", png, "Site1", 53,
                              page_rows=52, purge_from=10, gap_rows=1,
                              header_count=6)

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        assert ws.cell(row=53, column=1).value == "Site1"
        wb.close()

    def test_snap_with_headers_mid_page(self, tmp_path):
        """start_row=54, page_rows=52, header_count=6, purge_from=10 → label at row 99 (snapped)."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "test.png"
        _make_test_png(png, 10, 10)  # rows_needed = 1

        next_row = insert_png(output, "Sheet", png, "Site1", 54,
                              page_rows=52, purge_from=10, gap_rows=1,
                              header_count=6)

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        assert ws.cell(row=99, column=1).value == "Site1"
        wb.close()

    def test_snap_no_headers_unchanged(self, tmp_path):
        """start_row=53, page_rows=52, header_count=0, purge_from=10 → label at row 53 (original behavior)."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "test.png"
        _make_test_png(png, 10, 10)  # rows_needed = 1

        next_row = insert_png(output, "Sheet", png, "Site1", 53,
                              page_rows=52, purge_from=10, gap_rows=1,
                              header_count=0)

        wb = load_workbook(str(output))
        ws = wb["Sheet"]
        assert ws.cell(row=53, column=1).value == "Site1"
        wb.close()

    def test_overflow_no_label_with_headers(self, tmp_path):
        """No-label overflow guard respects header_count=2.
        header_count reduces content area, pushing image past page boundary."""
        output = _make_test_xlsx(tmp_path, "test.xlsx")
        png = tmp_path / "tall.png"
        _make_test_png(png, 10, 100)  # rows_needed = ceil(100*0.75/15) = 5

        # page_rows=10, header_count=2, start_row=9, gap_rows=0
        # content_rows = 10-2 = 8
        # img_end = 9+0+5 = 14
        # start_row(9) <= page_rows(10) → page_end = page_rows = 10
        # 14 > 10 → pushed to page_end+1 = 11
        # img_row = 11+0 = 11, return = 11+5+0 = 16
        next_row = insert_png_no_label(output, "Sheet", png, start_row=9,
                                        gap_rows=0, page_rows=10, header_count=2)
        assert next_row > 10  # pushed past page boundary
        assert next_row == 16

    def test_full_pipeline_with_headers(self, tmp_path):
        """Multi-sheet workbook with header_count=3 — snap + overflow + multi-sheet."""
        xlsx_path = _make_test_xlsx(tmp_path, "integration.xlsx")
        png1 = tmp_path / "test1.png"
        png2 = tmp_path / "test2.png"
        png3 = tmp_path / "test3.png"
        _make_test_png(png1, 10, 10)
        _make_test_png(png2, 10, 10)
        _make_test_png(png3, 10, 10)

        wb = load_workbook(str(xlsx_path))
        wb.create_sheet("Sheet2")
        wb.save(str(xlsx_path))
        wb.close()

        # Sheet 1: first site at purge_from → no snap (start_row == purge_from)
        r1 = insert_png(xlsx_path, "Sheet", png1, "SiteA", 5,
                         page_rows=10, header_count=3, purge_from=5, gap_rows=0)
        # start_row=5 not > purge_from=5 → no snap
        # img_end=5+1+0+1=7, page_end=10 → no overflow
        # label@5, img@6, next=6+1+0=7
        assert r1 == 7

        # Sheet 1: no-label insert following r1 → overflow check only (no snap in no-label)
        r2 = insert_png_no_label(xlsx_path, "Sheet", png2, r1,
                                  page_rows=10, header_count=3, gap_rows=0)
        # start_row=7, rows_needed=1 → img_end=7+0+1=8
        # start_row<=page_rows → page_end=10, 8>10? NO → no push
        # img@7, next=7+1+0=8
        assert r2 == 8

        # Sheet 1: second site after gap → snap fires (start_row > purge_from)
        r3 = insert_png(xlsx_path, "Sheet", png3, "SiteB", 12,
                         page_rows=10, header_count=3, purge_from=5, gap_rows=0)
        # snap: start_row=12>5, content_rows=7, offset=12-10-2=0, pages_after=1 → page_end=18
        # overflow: img_end=18+1+0+1=20, offset=7, pages_before=1 → page_end=24, 20>24? NO
        # label@18, img@19, next=19+1+0=20
        assert r3 == 20

        # Sheet 2: analogous first site → no snap
        r4 = insert_png(xlsx_path, "Sheet2", png1, "SiteC", 5,
                         page_rows=10, header_count=3, purge_from=5, gap_rows=0)
        assert r4 == 7

        # Verify label placements across sheets
        wb = load_workbook(str(xlsx_path))
        ws = wb["Sheet"]
        assert ws.cell(row=5, column=1).value == "SiteA"
        assert ws.cell(row=18, column=1).value == "SiteB"
        ws2 = wb["Sheet2"]
        assert ws2.cell(row=5, column=1).value == "SiteC"
        wb.close()
