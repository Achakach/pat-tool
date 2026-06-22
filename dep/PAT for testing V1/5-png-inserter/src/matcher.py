from pathlib import Path
from openpyxl import load_workbook


def read_matching(file_path: str, sheet_name: str, filename_col: str, planwork_col: str) -> dict[str, list[str]]:
    """Read matching XLSX. Row 1 = headers, data from row 2.
    Returns {filename: [planwork, ...]}. Blank filename cells inherit from the row above."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # Row 1: find columns by header text
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column

    fn_idx = headers.get(filename_col.lower())
    pw_idx = headers.get(planwork_col.lower())

    if fn_idx is None:
        raise ValueError(f"Column '{filename_col}' not found in headers (row 1)")
    if pw_idx is None:
        raise ValueError(f"Column '{planwork_col}' not found in headers (row 1)")

    result = {}
    current_filename = None

    for row in ws.iter_rows(min_row=2):  # skip header row
        fn_cell = row[fn_idx - 1] if len(row) >= fn_idx else None
        pw_cell = row[pw_idx - 1] if len(row) >= pw_idx else None

        filename = str(fn_cell.value).strip() if fn_cell and fn_cell.value else None
        planwork = str(pw_cell.value).strip() if pw_cell and pw_cell.value else None

        # Inherit filename from above if blank
        if filename:
            current_filename = filename
        if not current_filename:
            continue
        if not planwork:
            continue

        if current_filename not in result:
            result[current_filename] = []
        result[current_filename].append(planwork)

    wb.close()
    return result


def extract_planwork(filename: str) -> str:
    """Extract planwork prefix from PNG filename.
    'PW planwork100_exist BKK01_Bayface Before.png' -> 'PW planwork100'"""
    stem = Path(filename).stem
    parts = stem.split("_", 1)
    return parts[0] if parts else stem


def match_pngs(png_folder: Path, planworks: list[str]) -> list[Path]:
    """Return PNG files whose planwork prefix matches any of the given planworks."""
    matches = []
    for png in sorted(png_folder.glob("*.png")):
        pw = extract_planwork(png.name)
        for planwork in planworks:
            if planwork in pw:
                matches.append(png)
                break
    return matches
