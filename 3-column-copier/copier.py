#!/usr/bin/env python3
"""Column copier — create temp columns, copy to target, cleanup."""

import sys
import json
import shutil
import re
from pathlib import Path
from openpyxl import load_workbook
from src.columns import (
    col_letter_to_index, build_pw_column, build_ip_column,
    copy_column, find_matching_sheet, clean_sheet_name
)
from src.print_setup import _setup_a4_print, _calc_page_rows, _parse_print_title_rows, snap_gap_rows


def read_matching(file_path, sheet_name, filename_col, planwork_col):
    wb = load_workbook(str(file_path), data_only=True)
    ws = wb[sheet_name]
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column
    fn_idx = headers.get(filename_col.lower())
    pw_idx = headers.get(planwork_col.lower())
    if not fn_idx or not pw_idx:
        wb.close()
        return {}
    result = {}
    current_filename = None
    for row in ws.iter_rows(min_row=2):
        fn = str(row[fn_idx - 1].value).strip() if row[fn_idx - 1].value else None
        pw = str(row[pw_idx - 1].value).strip() if row[pw_idx - 1].value else None
        if fn:
            current_filename = fn
        if current_filename and pw:
            result[pw] = current_filename
    wb.close()
    return result


def main():
    config_path = Path(__file__).parent / "config.json"
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    action = config.get("action", "copy")
    matching_file = (Path(__file__).parent / config["matching_file"]).resolve()
    source_folder = (Path(__file__).parent / config["source_folder"]).resolve()
    target_folder = (Path(__file__).parent / config["target_folder"]).resolve()
    output_folder = (Path(__file__).parent / config["output_folder"]).resolve()

    # Read matching
    pw_to_file = read_matching(
        str(matching_file), config["matching_sheet"],
        config["filename_col"], config["planwork_col"]
    )

    data_sheet = config["data_sheet"]
    target_sheet_name = config["target_sheet"]
    start_row = config["source_start_row"]
    paste_row = config["paste_start_row"]
    paste_mode = config.get("paste_mode", "overwrite")
    columns = config["columns"]

    # Parse print config
    print_title_rows_raw = config.get("print_title_rows")
    header_count, print_title_rows_str = _parse_print_title_rows(print_title_rows_raw)
    page_break_enabled = config.get("page_break_enabled", False)

    output_folder.mkdir(parents=True, exist_ok=True)

    if action == "copy":
        for xlsx_path in sorted(source_folder.glob("*.xlsx")):
            paste_row = config["paste_start_row"]  # reset for each file
            if xlsx_path.name.startswith("~$"):
                continue
            print(f"Processing: {xlsx_path.name}")

            # Open source
            wb = load_workbook(str(xlsx_path))

            # Step 1: Find PW sheet -> extract planwork
            planwork = None
            for sn in wb.sheetnames:
                m = re.match(r'^PW\s+(.+)', sn, re.IGNORECASE)
                if m:
                    planwork = m.group(1).strip()
                    break
            if not planwork:
                print(f"  SKIP: No PW sheet found")
                wb.close()
                continue

            # Step 2: Build temp columns in source
            ws = wb[data_sheet]
            for col_name, col_cfg in columns.items():
                col_type = col_cfg.get("type")
                paste_to = col_cfg.get("paste_to")
                if col_type == "planwork":
                    build_pw_column(ws, planwork, paste_to, start_row)
                    print(f"  PW column: {paste_to} = '{planwork}'")
                elif col_type == "ip_lookup":
                    log_sheet_name = col_cfg["log_sheet"]
                    if log_sheet_name in wb.sheetnames:
                        log_ws = wb[log_sheet_name]
                        build_ip_column(ws, col_cfg["lookup_col"], log_ws, paste_to, start_row)
                        print(f"  IP column: {paste_to} (lookup from {col_cfg['lookup_col']})")

            wb.save(str(xlsx_path))
            wb.close()

            # Step 3: Find target via planwork matching
            target_file = pw_to_file.get(planwork)
            if not target_file:
                print(f"  SKIP: No target for planwork '{planwork}'")
                continue
            if not target_file.endswith(".xlsx"):
                target_file += ".xlsx"

            target_path = target_folder / target_file
            if not target_path.exists():
                print(f"  SKIP: Target not found: {target_file}")
                continue

            # Step 4: Open target, copy columns
            twb = load_workbook(str(target_path))
            tsheet_name = find_matching_sheet(twb, target_sheet_name)
            if not tsheet_name:
                print(f"  SKIP: Target sheet '{target_sheet_name}' not found")
                twb.close()
                continue

            tws = twb[tsheet_name]

            _setup_a4_print(tws, print_title_rows_str)
            if page_break_enabled:
                page_rows = _calc_page_rows(tws, config.get("a4_page_rows"))
            else:
                page_rows = None

            swb = load_workbook(str(xlsx_path))
            sws = swb[data_sheet]

            # If append mode, find first blank row in target sheet
            if paste_mode == "append":
                actual_row = paste_row
                max_row_check = tws.max_row + 100
                while actual_row < max_row_check:
                    empty = True
                    for c in range(1, tws.max_column + 1):
                        if tws.cell(row=actual_row, column=c).value is not None:
                            empty = False
                            break
                    if empty:
                        break
                    actual_row += 1
                paste_row = actual_row

            # Append mode: insert rows to push existing content down
            if paste_mode == "append" and page_break_enabled:
                # Count source data rows
                src_data_rows = 0
                check_row = start_row
                while True:
                    empty = all(sws.cell(row=check_row, column=c).value is None
                                for c in range(1, sws.max_column + 1))
                    if empty and check_row > start_row:
                        break
                    src_data_rows += 1
                    check_row += 1

                # Check for merged cells in paste range
                has_merged = False
                for merged_range in tws.merged_cells.ranges:
                    if merged_range.min_row <= paste_row + src_data_rows and merged_range.max_row >= paste_row:
                        has_merged = True
                        break

                if has_merged:
                    print(f"WARNING: Merged cells detected in paste area, skipping insert_rows", file=sys.stderr)
                elif src_data_rows > 0:
                    tws.insert_rows(paste_row, src_data_rows)

            paste_end = paste_row  # track max row reached across all columns
            for col_name, col_cfg in columns.items():
                col_type = col_cfg.get("type")
                if col_type == "copy":
                    copy_column(sws, col_cfg["source_col"], col_cfg["paste_to"], start_row)

                paste_to = col_cfg.get("paste_to")
                # Copy value from source to target
                src_idx = col_letter_to_index(paste_to)
                dst_idx = col_letter_to_index(paste_to)
                src_row = start_row
                dst_row = paste_row if paste_mode == "append" else min(start_row, paste_row)
                while True:
                    row_empty = True
                    for c in range(1, sws.max_column + 1):
                        if sws.cell(row=src_row, column=c).value is not None:
                            row_empty = False
                            break
                    if row_empty and src_row > start_row:
                        break
                    val = sws.cell(row=src_row, column=src_idx).value
                    if val is not None:
                        tws.cell(row=dst_row, column=dst_idx).value = val
                    src_row += 1
                    dst_row += 1
                paste_end = max(paste_end, dst_row)

            # Page-overflow snap: push content below to clean page boundary
            if page_break_enabled:
                gap = snap_gap_rows(paste_end, tws, page_rows, header_count)
                if gap > 0:
                    tws.insert_rows(paste_end, gap)
                    print(f"  Snapped: inserted {gap} gap rows at row {paste_end} to push content to page boundary", file=sys.stderr)

            swb.close()

            # Save target to output
            out_path = output_folder / target_file
            twb.save(str(out_path))
            twb.close()
            print(f"  Copied to: {target_file}")

        print(f"\nDone.")

    elif action == "cleanup":
        # Delete PW and IP columns from source
        for xlsx_path in sorted(source_folder.glob("*.xlsx")):
            if xlsx_path.name.startswith("~$"):
                continue
            wb = load_workbook(str(xlsx_path))
            ws = wb[data_sheet]
            for col_name, col_cfg in columns.items():
                col_type = col_cfg.get("type")
                if col_type in ("planwork", "ip_lookup"):
                    paste_to = col_cfg.get("paste_to")
                    col_idx = col_letter_to_index(paste_to)
                    ws.delete_cols(col_idx)
            out = output_folder / xlsx_path.name
            wb.save(str(out))
            wb.close()
        print("Cleanup done.")


if __name__ == "__main__":
    main()
