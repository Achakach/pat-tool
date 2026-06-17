import re
from pathlib import Path
from openpyxl import load_workbook


def col_letter_to_index(letter: str) -> int:
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def build_pw_column(ws, planwork: str, col_letter: str, start_row: int = 2, lookup_col: str = None):
    """Fill rows in the column with planwork value. If lookup_col is provided, stop when that column has no data; otherwise stop when row is fully empty."""
    col_idx = col_letter_to_index(col_letter)
    lookup_idx = col_letter_to_index(lookup_col) if lookup_col else None
    row = start_row
    while True:
        if lookup_idx:
            # Stop when lookup column (NE_NO) is empty
            if ws.cell(row=row, column=lookup_idx).value is None:
                if row > start_row:
                    break
        else:
            # Original: stop when ALL columns empty
            row_empty = True
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=c).value is not None:
                    row_empty = False
                    break
            if row_empty and row > start_row:
                break
        ws.cell(row=row, column=col_idx).value = planwork
        row += 1


def build_ip_column(ws, lookup_col: str, log_sheet, col_letter: str, start_row: int = 2):
    """Lookup NE_NO values in log sheet row 1, fill IP column."""
    lookup_idx = col_letter_to_index(lookup_col)
    col_idx = col_letter_to_index(col_letter)

    ip_map = {}
    for cell in log_sheet[1]:
        if cell.value:
            text = str(cell.value)
            match = re.search(r'([^_]+)_(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', text)
            if match:
                ip_map[match.group(1).strip()] = match.group(2).strip()

    row = start_row
    while True:
        row_empty = True
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=c).value is not None:
                row_empty = False
                break
        if row_empty and row > start_row:
            break

        ne_no = ws.cell(row=row, column=lookup_idx).value
        if ne_no:
            ip = ip_map.get(str(ne_no).strip(), "")
            ws.cell(row=row, column=col_idx).value = ip
        row += 1



def clean_sheet_name(name: str) -> str:
    """Clean sheet name for matching. '2.3. IP & Port Assignment(P.4)' -> 'ip & port assignment'"""
    name = re.sub(r"^\d+\.?\d*\.?\s*", "", name)
    name = re.sub(r"\([^)]*\)", "", name)
    name = name.replace("_", " ")
    return " ".join(name.split()).lower()


def find_matching_sheet(wb, target_name: str) -> str | None:
    """Find sheet matching target name (with noise stripping)."""
    clean_target = clean_sheet_name(target_name)
    for sheet_name in wb.sheetnames:
        if clean_sheet_name(sheet_name) == clean_target:
            return sheet_name
    return None
